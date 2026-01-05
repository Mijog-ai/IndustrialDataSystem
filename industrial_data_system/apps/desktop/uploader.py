"""Enhanced PyQt5 UI with test type organization for uploads."""

from __future__ import annotations

import csv
import io
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from dotenv import load_dotenv
from PyQt5.QtCore import QSize, Qt, QUrl, pyqtSignal
from PyQt5.QtGui import QColor, QDesktopServices, QFont, QIcon, QKeySequence, QPalette
from PyQt5.QtWidgets import (
    QAction,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFrame,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMenuBar,
    QMessageBox,
    QProgressDialog,
    QPushButton,
    QScrollArea,
    QStackedWidget,
    QStatusBar,
    QTableWidget,
    QTableWidgetItem,
    QToolBar,
    QVBoxLayout,
    QWidget,
)

from industrial_data_system.core.auth import LocalAuthStore, LocalUser, UploadHistoryStore
from industrial_data_system.core.config import get_config
from industrial_data_system.core.constants import MAX_PREVIEW_ROWS, SUPPORTED_EXTENSIONS
from industrial_data_system.core.db_manager import DatabaseManager
from industrial_data_system.core.storage import LocalStorageManager, StorageError

# ---------------------------------------------------------------------------
# Environment loading helpers
# ---------------------------------------------------------------------------


def _load_environment() -> None:
    """Load environment variables with support for frozen executables."""

    candidate_paths = []

    # 1. Directory that contains the running script (useful in development).
    script_directory = Path(__file__).resolve().parent
    candidate_paths.append(script_directory / ".env")
    candidate_paths.append(script_directory.parent / ".env")
    candidate_paths.append(script_directory.parent.parent / ".env")

    # 2. When running from a PyInstaller bundle, assets live under ``_MEIPASS``.
    meipass_dir = getattr(sys, "_MEIPASS", None)
    if meipass_dir:
        candidate_paths.append(Path(meipass_dir) / ".env")

    # 3. Current working directory (covers the case where the .exe is launched
    #    from a shortcut or another folder but the .env sits next to it).
    candidate_paths.append(Path.cwd() / ".env")

    for env_path in candidate_paths:
        if env_path.is_file():
            load_dotenv(env_path)
            break
    else:
        # Fall back to the default search behaviour which looks up the tree.
        load_dotenv()


_load_environment()

CONFIG = get_config()


class DesktopTheme:
    """Traditional desktop application design system."""

    # Traditional Desktop Color Palette
    PRIMARY = "#0078D4"  # Windows Blue
    PRIMARY_LIGHT = "#429CE3"
    PRIMARY_DARK = "#005A9E"

    SECONDARY = "#767676"  # Medium Gray
    SECONDARY_LIGHT = "#A0A0A0"

    SUCCESS = "#107C10"  # Dark Green
    WARNING = "#FF8C00"  # Dark Orange
    ERROR = "#D13438"  # Dark Red

    BACKGROUND = "#F0F0F0"  # Classic Gray
    SURFACE = "#FFFFFF"  # White
    SURFACE_DARK = "#E6E6E6"

    GROUPBOX_BG = "#FAFAFA"  # Light background for group boxes

    TEXT_PRIMARY = "#000000"  # Black
    TEXT_SECONDARY = "#666666"  # Dark Gray
    TEXT_HINT = "#999999"  # Light Gray

    BORDER = "#ABABAB"
    BORDER_DARK = "#707070"
    BORDER_FOCUS = "#0078D4"

    @staticmethod
    def get_stylesheet():
        """Return complete application stylesheet with traditional desktop styling."""
        return f"""
            QMainWindow {{
                background-color: {DesktopTheme.BACKGROUND};
            }}

            QWidget {{
                font-family: 'Segoe UI', 'Tahoma', 'MS Sans Serif', Arial, sans-serif;
                font-size: 9pt;
                color: {DesktopTheme.TEXT_PRIMARY};
            }}

            QMenuBar {{
                background-color: {DesktopTheme.SURFACE};
                border-bottom: 1px solid {DesktopTheme.BORDER};
                padding: 2px;
            }}

            QMenuBar::item {{
                background-color: transparent;
                padding: 4px 8px;
            }}

            QMenuBar::item:selected {{
                background-color: {DesktopTheme.PRIMARY_LIGHT};
                color: white;
            }}

            QMenu {{
                background-color: {DesktopTheme.SURFACE};
                border: 1px solid {DesktopTheme.BORDER};
            }}

            QMenu::item {{
                padding: 4px 20px 4px 8px;
            }}

            QMenu::item:selected {{
                background-color: {DesktopTheme.PRIMARY_LIGHT};
                color: white;
            }}

            QToolBar {{
                background-color: {DesktopTheme.SURFACE};
                border-bottom: 1px solid {DesktopTheme.BORDER};
                spacing: 2px;
                padding: 2px;
            }}

            QToolButton {{
                background-color: transparent;
                border: 1px solid transparent;
                padding: 3px;
                margin: 1px;
            }}

            QToolButton:hover {{
                background-color: {DesktopTheme.PRIMARY_LIGHT};
                border: 1px solid {DesktopTheme.PRIMARY};
            }}

            QToolButton:pressed {{
                background-color: {DesktopTheme.PRIMARY_DARK};
                border: 1px solid {DesktopTheme.PRIMARY_DARK};
            }}

            QStatusBar {{
                background-color: {DesktopTheme.SURFACE};
                border-top: 1px solid {DesktopTheme.BORDER};
            }}

            QLabel {{
                color: {DesktopTheme.TEXT_PRIMARY};
                background-color: transparent;
            }}

            QLabel[heading="true"] {{
                font-size: 12pt;
                font-weight: bold;
                color: {DesktopTheme.TEXT_PRIMARY};
                padding: 4px 0px;
            }}

            QLabel[subheading="true"] {{
                font-size: 10pt;
                font-weight: bold;
                color: {DesktopTheme.TEXT_PRIMARY};
                padding: 2px 0px;
            }}

            QLabel[caption="true"] {{
                font-size: 8pt;
                color: {DesktopTheme.TEXT_SECONDARY};
            }}

            QLineEdit {{
                padding: 3px 5px;
                border: 1px solid {DesktopTheme.BORDER};
                background-color: {DesktopTheme.SURFACE};
                color: {DesktopTheme.TEXT_PRIMARY};
                selection-background-color: {DesktopTheme.PRIMARY};
            }}

            QLineEdit:focus {{
                border: 1px solid {DesktopTheme.BORDER_FOCUS};
            }}

            QLineEdit:disabled {{
                background-color: {DesktopTheme.SURFACE_DARK};
                color: {DesktopTheme.TEXT_HINT};
            }}

            QComboBox {{
                padding: 3px 5px;
                border: 1px solid {DesktopTheme.BORDER};
                background-color: {DesktopTheme.SURFACE};
                color: {DesktopTheme.TEXT_PRIMARY};
                min-height: 20px;
            }}

            QComboBox:focus {{
                border: 1px solid {DesktopTheme.BORDER_FOCUS};
            }}

            QComboBox::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 18px;
                border-left: 1px solid {DesktopTheme.BORDER};
            }}

            QComboBox::down-arrow {{
                image: none;
                border: 2px solid {DesktopTheme.TEXT_PRIMARY};
                width: 6px;
                height: 6px;
                border-top: none;
                border-right: none;
                transform: rotate(-45deg);
            }}

            QComboBox QAbstractItemView {{
                background-color: {DesktopTheme.SURFACE};
                border: 1px solid {DesktopTheme.BORDER};
                selection-background-color: {DesktopTheme.PRIMARY};
                selection-color: white;
                outline: none;
            }}

            QPushButton {{
                padding: 4px 12px;
                border: 1px solid {DesktopTheme.BORDER};
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {DesktopTheme.SURFACE}, stop:1 {DesktopTheme.SURFACE_DARK});
                color: {DesktopTheme.TEXT_PRIMARY};
                min-height: 22px;
            }}

            QPushButton:hover {{
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #E5F3FF, stop:1 #D0E8FF);
                border: 1px solid {DesktopTheme.PRIMARY};
            }}

            QPushButton:pressed {{
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {DesktopTheme.SURFACE_DARK}, stop:1 {DesktopTheme.SURFACE});
                border: 1px solid {DesktopTheme.PRIMARY_DARK};
            }}

            QPushButton[primary="true"] {{
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {DesktopTheme.PRIMARY_LIGHT}, stop:1 {DesktopTheme.PRIMARY});
                color: white;
                border: 1px solid {DesktopTheme.PRIMARY_DARK};
                font-weight: bold;
            }}

            QPushButton[primary="true"]:hover {{
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {DesktopTheme.PRIMARY}, stop:1 {DesktopTheme.PRIMARY_LIGHT});
            }}

            QPushButton[primary="true"]:pressed {{
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {DesktopTheme.PRIMARY_DARK}, stop:1 {DesktopTheme.PRIMARY});
            }}

            QPushButton[secondary="true"] {{
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {DesktopTheme.SURFACE}, stop:1 {DesktopTheme.SURFACE_DARK});
                color: {DesktopTheme.TEXT_PRIMARY};
                border: 1px solid {DesktopTheme.BORDER};
            }}

            QPushButton[danger="true"] {{
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #F44336, stop:1 {DesktopTheme.ERROR});
                color: white;
                border: 1px solid #C62828;
                font-weight: bold;
            }}

            QPushButton[danger="true"]:hover {{
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #FF5252, stop:1 #F44336);
            }}

            QPushButton[flat="true"] {{
                background-color: transparent;
                color: {DesktopTheme.PRIMARY};
                border: none;
                padding: 4px 8px;
            }}

            QPushButton[flat="true"]:hover {{
                background-color: #E5F3FF;
            }}

            QPushButton:disabled {{
                background-color: {DesktopTheme.SURFACE_DARK};
                color: {DesktopTheme.TEXT_HINT};
                border: 1px solid {DesktopTheme.BORDER};
            }}

            QTableWidget {{
                background-color: {DesktopTheme.SURFACE};
                border: 1px solid {DesktopTheme.BORDER};
                gridline-color: {DesktopTheme.BORDER};
                selection-background-color: {DesktopTheme.PRIMARY};
                selection-color: white;
            }}

            QTableWidget::item {{
                padding: 4px;
            }}

            QHeaderView::section {{
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {DesktopTheme.SURFACE}, stop:1 {DesktopTheme.SURFACE_DARK});
                padding: 4px;
                border: 1px solid {DesktopTheme.BORDER};
                font-weight: bold;
                color: {DesktopTheme.TEXT_PRIMARY};
            }}

            QGroupBox {{
                border: 1px solid {DesktopTheme.BORDER};
                margin-top: 8px;
                padding-top: 8px;
                background-color: {DesktopTheme.GROUPBOX_BG};
            }}

            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 8px;
                padding: 0 4px;
                background-color: {DesktopTheme.GROUPBOX_BG};
                color: {DesktopTheme.TEXT_PRIMARY};
                font-weight: bold;
            }}

            QFrame[card="true"] {{
                background-color: {DesktopTheme.GROUPBOX_BG};
                border: 1px solid {DesktopTheme.BORDER};
                padding: 8px;
            }}

            QDialog {{
                background-color: {DesktopTheme.BACKGROUND};
            }}

            QScrollBar:vertical {{
                border: 1px solid {DesktopTheme.BORDER};
                background-color: {DesktopTheme.SURFACE_DARK};
                width: 16px;
            }}

            QScrollBar::handle:vertical {{
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {DesktopTheme.SURFACE}, stop:1 {DesktopTheme.SURFACE_DARK});
                border: 1px solid {DesktopTheme.BORDER};
                min-height: 20px;
            }}

            QScrollBar::handle:vertical:hover {{
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #E5F3FF, stop:1 #D0E8FF);
            }}

            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                border: 1px solid {DesktopTheme.BORDER};
                background-color: {DesktopTheme.SURFACE_DARK};
                height: 16px;
                subcontrol-origin: margin;
            }}

            QScrollBar::add-line:vertical {{
                subcontrol-position: bottom;
            }}

            QScrollBar::sub-line:vertical {{
                subcontrol-position: top;
            }}

            QScrollBar:horizontal {{
                border: 1px solid {DesktopTheme.BORDER};
                background-color: {DesktopTheme.SURFACE_DARK};
                height: 16px;
            }}

            QScrollBar::handle:horizontal {{
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {DesktopTheme.SURFACE}, stop:1 {DesktopTheme.SURFACE_DARK});
                border: 1px solid {DesktopTheme.BORDER};
                min-width: 20px;
            }}

            QCheckBox::indicator {{
                width: 13px;
                height: 13px;
                border: 1px solid {DesktopTheme.BORDER_DARK};
                background-color: {DesktopTheme.SURFACE};
            }}

            QCheckBox::indicator:hover {{
                border-color: {DesktopTheme.PRIMARY};
                background-color: #E5F3FF;
            }}

            QCheckBox::indicator:checked {{
                background-color: {DesktopTheme.PRIMARY};
                border-color: {DesktopTheme.PRIMARY_DARK};
            }}

            QTabWidget::pane {{
                border: 1px solid {DesktopTheme.BORDER};
                background-color: {DesktopTheme.SURFACE};
                padding: 4px;
            }}

            QTabBar::tab {{
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {DesktopTheme.SURFACE}, stop:1 {DesktopTheme.SURFACE_DARK});
                color: {DesktopTheme.TEXT_PRIMARY};
                border: 1px solid {DesktopTheme.BORDER};
                border-bottom: none;
                padding: 4px 12px;
                margin-right: 2px;
                min-width: 80px;
            }}

            QTabBar::tab:selected {{
                background-color: {DesktopTheme.SURFACE};
                color: {DesktopTheme.TEXT_PRIMARY};
                font-weight: bold;
                border-bottom: 1px solid {DesktopTheme.SURFACE};
            }}

            QTabBar::tab:hover:!selected {{
                background-color: #E5F3FF;
            }}
        """


class NewPumpSeriesDialog(QDialog):
    """Dialog for creating a new pump series."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Create New Pump Series")
        self.setMinimumWidth(400)

        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(24, 24, 24, 24)

        title = QLabel("New Pump Series")
        title.setProperty("subheading", True)
        layout.addWidget(title)

        desc = QLabel(
            "Enter a name for the new pump series. A folder hierarchy will be created on the shared drive."
        )
        desc.setProperty("caption", True)
        desc.setWordWrap(True)
        layout.addWidget(desc)

        input_label = QLabel("Pump Series Name")
        input_label.setStyleSheet(f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;")
        layout.addWidget(input_label)

        self.series_input = QLineEdit()
        self.series_input.setPlaceholderText("e.g., Alpha Series, Beta Series")
        layout.addWidget(self.series_input)

        description_label = QLabel("Description (optional)")
        description_label.setStyleSheet(
            f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;"
        )
        layout.addWidget(description_label)

        self.description_input = QLineEdit()
        self.description_input.setPlaceholderText("Short summary for this pump series")
        layout.addWidget(self.description_input)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)

        for button in button_box.buttons():
            if button_box.buttonRole(button) == QDialogButtonBox.AcceptRole:
                button.setProperty("primary", True)
            else:
                button.setProperty("secondary", True)
            button.setMinimumHeight(44)

        layout.addWidget(button_box)

    def get_pump_series(self) -> str:
        return self.series_input.text().strip()

    def get_description(self) -> str:
        return self.description_input.text().strip()


class NewTestTypeDialog(QDialog):
    """Dialog for creating a new test type."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Create New Test Type")
        self.setMinimumWidth(400)

        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(24, 24, 24, 24)

        # Title
        title = QLabel("New Test Type")
        title.setProperty("subheading", True)
        layout.addWidget(title)

        # Description
        desc = QLabel(
            "Enter a name for the new test type. A matching folder will be created on the shared drive."
        )
        desc.setProperty("caption", True)
        desc.setWordWrap(True)
        layout.addWidget(desc)

        # Input field
        input_label = QLabel("Test Type Name")
        input_label.setStyleSheet(f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;")
        layout.addWidget(input_label)

        self.test_type_input = QLineEdit()
        self.test_type_input.setPlaceholderText("e.g., Performance Test, Load Test, Stress Test")
        layout.addWidget(self.test_type_input)

        description_label = QLabel("Description (optional)")
        description_label.setStyleSheet(
            f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;"
        )
        layout.addWidget(description_label)

        self.description_input = QLineEdit()
        self.description_input.setPlaceholderText("Short summary for this test type")
        layout.addWidget(self.description_input)

        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)

        # Style buttons
        for button in button_box.buttons():
            if button_box.buttonRole(button) == QDialogButtonBox.AcceptRole:
                button.setProperty("primary", True)
            else:
                button.setProperty("secondary", True)
            button.setMinimumHeight(44)

        layout.addWidget(button_box)

    def get_test_type(self) -> str:
        return self.test_type_input.text().strip()

    def get_description(self) -> str:
        return self.description_input.text().strip()


class SessionState:
    """Minimal session holder for the desktop application."""

    def __init__(self) -> None:
        self.user: Optional[Dict[str, Any]] = None

    def set_user(self, user: Dict[str, Any]) -> None:
        self.user = user

    def clear(self) -> None:
        self.user = None


class LoginPage(QWidget):
    """Modern authentication interface with industrial design."""

    login_requested = pyqtSignal(str, str)
    signup_requested = pyqtSignal(str, str, str)
    forgot_password_requested = pyqtSignal()

    def __init__(self) -> None:
        super().__init__()

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setAlignment(Qt.AlignCenter)
        layout.setContentsMargins(40, 60, 40, 60)
        layout.setSpacing(32)

        # Header
        header_widget = QWidget()
        header_layout = QVBoxLayout(header_widget)
        header_layout.setSpacing(8)
        header_layout.setAlignment(Qt.AlignCenter)

        title = QLabel("Inline Data System")
        title.setProperty("heading", True)
        header_layout.addWidget(title, alignment=Qt.AlignCenter)

        subtitle = QLabel("Secure data management platform")
        subtitle.setProperty("caption", True)
        header_layout.addWidget(subtitle, alignment=Qt.AlignCenter)

        layout.addWidget(header_widget)

        # Tab Toggle
        toggle_card = QFrame()
        toggle_card.setProperty("card", True)
        toggle_card.setMaximumWidth(480)
        toggle_layout = QVBoxLayout(toggle_card)
        toggle_layout.setContentsMargins(8, 8, 8, 8)
        toggle_layout.setSpacing(0)

        tab_row = QHBoxLayout()
        tab_row.setSpacing(0)
        tab_row.setContentsMargins(0, 0, 0, 0)

        self.login_toggle = QPushButton("Sign In")
        self.signup_toggle = QPushButton("Sign Up")

        for button in (self.login_toggle, self.signup_toggle):
            button.setCheckable(True)
            button.setCursor(Qt.PointingHandCursor)
            button.setMinimumHeight(48)
            button.setProperty("secondary", True)

        self.login_toggle.clicked.connect(lambda: self._switch_mode("login"))
        self.signup_toggle.clicked.connect(lambda: self._switch_mode("signup"))

        tab_row.addWidget(self.login_toggle)
        tab_row.addWidget(self.signup_toggle)
        toggle_layout.addLayout(tab_row)

        layout.addWidget(toggle_card, alignment=Qt.AlignCenter)

        # Form Container
        form_card = QFrame()
        form_card.setProperty("card", True)
        form_card.setMaximumWidth(480)
        form_layout = QVBoxLayout(form_card)
        form_layout.setSpacing(24)
        form_layout.setContentsMargins(32, 32, 32, 32)

        self.form_stack = QStackedWidget()
        form_layout.addWidget(self.form_stack)

        # Login Form
        self.login_form = QWidget()
        login_layout = QVBoxLayout(self.login_form)
        login_layout.setSpacing(20)
        login_layout.setContentsMargins(0, 0, 0, 0)

        login_email_label = QLabel("Email or Username")
        login_email_label.setStyleSheet(
            f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;"
        )
        self.login_email_input = QLineEdit()
        self.login_email_input.setPlaceholderText("Enter your email or username")
        login_layout.addWidget(login_email_label)
        login_layout.addWidget(self.login_email_input)

        login_password_label = QLabel("Password")
        login_password_label.setStyleSheet(
            f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;"
        )
        self.login_password_input = QLineEdit()
        self.login_password_input.setPlaceholderText("Enter your password")
        self.login_password_input.setEchoMode(QLineEdit.Password)
        login_layout.addWidget(login_password_label)
        login_layout.addWidget(self.login_password_input)

        login_button = QPushButton("Sign In")
        login_button.setProperty("primary", True)
        login_button.clicked.connect(self._emit_login_request)
        login_layout.addWidget(login_button)

        forgot_button = QPushButton("Forgot Password?")
        forgot_button.setProperty("flat", True)
        forgot_button.setCursor(Qt.PointingHandCursor)
        forgot_button.clicked.connect(self.forgot_password_requested)
        login_layout.addWidget(forgot_button, alignment=Qt.AlignCenter)

        self.form_stack.addWidget(self.login_form)

        # Signup Form
        self.signup_form = QWidget()
        signup_layout = QVBoxLayout(self.signup_form)
        signup_layout.setSpacing(20)
        signup_layout.setContentsMargins(0, 0, 0, 0)

        signup_email_label = QLabel("Email")
        signup_email_label.setStyleSheet(
            f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;"
        )
        self.signup_email_input = QLineEdit()
        self.signup_email_input.setPlaceholderText("your.email@company.com")
        signup_layout.addWidget(signup_email_label)
        signup_layout.addWidget(self.signup_email_input)

        signup_username_label = QLabel("Username")
        signup_username_label.setStyleSheet(
            f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;"
        )
        self.signup_username_input = QLineEdit()
        self.signup_username_input.setPlaceholderText("Username (max 6 characters)")
        self.signup_username_input.setMaxLength(6)
        signup_layout.addWidget(signup_username_label)
        signup_layout.addWidget(self.signup_username_input)

        signup_password_label = QLabel("Password")
        signup_password_label.setStyleSheet(
            f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;"
        )
        self.signup_password_input = QLineEdit()
        self.signup_password_input.setPlaceholderText("Password (max 6 characters)")
        self.signup_password_input.setEchoMode(QLineEdit.Password)
        self.signup_password_input.setMaxLength(6)
        signup_layout.addWidget(signup_password_label)
        signup_layout.addWidget(self.signup_password_input)

        signup_confirm_label = QLabel("Confirm Password")
        signup_confirm_label.setStyleSheet(
            f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;"
        )
        self.signup_confirm_input = QLineEdit()
        self.signup_confirm_input.setPlaceholderText("Re-enter your password")
        self.signup_confirm_input.setEchoMode(QLineEdit.Password)
        self.signup_confirm_input.setMaxLength(6)
        signup_layout.addWidget(signup_confirm_label)
        signup_layout.addWidget(self.signup_confirm_input)

        signup_button = QPushButton("Create Account")
        signup_button.setProperty("primary", True)
        signup_button.clicked.connect(self._emit_signup_request)
        signup_layout.addWidget(signup_button)

        self.form_stack.addWidget(self.signup_form)

        layout.addWidget(form_card, alignment=Qt.AlignCenter)
        layout.addStretch()

        scroll.setWidget(container)
        main_layout.addWidget(scroll)

        self._switch_mode("login")

    def show_login(self, email: str = "") -> None:
        self.login_email_input.setText(email)
        self.login_password_input.clear()
        self.signup_email_input.clear()
        self.signup_username_input.clear()
        self.signup_password_input.clear()
        self.signup_confirm_input.clear()
        self._switch_mode("login")
        if email:
            self.login_password_input.setFocus()
        else:
            self.login_email_input.setFocus()

    def show_signup(self) -> None:
        self.login_email_input.clear()
        self.login_password_input.clear()
        self._switch_mode("signup")
        self.signup_email_input.setFocus()

    def _switch_mode(self, mode: str) -> None:
        if mode == "signup":
            self.form_stack.setCurrentWidget(self.signup_form)
            self.signup_toggle.setChecked(True)
            self.login_toggle.setChecked(False)
            self.signup_toggle.setProperty("primary", True)
            self.signup_toggle.setProperty("secondary", False)
            self.login_toggle.setProperty("primary", False)
            self.login_toggle.setProperty("secondary", True)
        else:
            self.form_stack.setCurrentWidget(self.login_form)
            self.login_toggle.setChecked(True)
            self.signup_toggle.setChecked(False)
            self.login_toggle.setProperty("primary", True)
            self.login_toggle.setProperty("secondary", False)
            self.signup_toggle.setProperty("primary", False)
            self.signup_toggle.setProperty("secondary", True)

        self.login_toggle.style().unpolish(self.login_toggle)
        self.login_toggle.style().polish(self.login_toggle)
        self.signup_toggle.style().unpolish(self.signup_toggle)
        self.signup_toggle.style().polish(self.signup_toggle)

    def _emit_login_request(self) -> None:
        identifier = self.login_email_input.text().strip()
        password = self.login_password_input.text()
        if not identifier or not password:
            QMessageBox.warning(
                self,
                "Inline Data System",
                "Username/Email and password are required.",
            )
            return
        self.login_requested.emit(identifier, password)

    def _emit_signup_request(self) -> None:
        email = self.signup_email_input.text().strip()
        username = self.signup_username_input.text().strip()
        password = self.signup_password_input.text()
        confirm = self.signup_confirm_input.text()

        if not email or not username or not password or not confirm:
            QMessageBox.warning(self, "Inline Data System", "All signup fields are required.")
            return

        if password != confirm:
            QMessageBox.warning(self, "Inline Data System", "Passwords do not match.")
            return

        self.signup_requested.emit(email, username, password)


class ForgotPasswordPage(QWidget):
    """Password reset interface with modern design."""

    reset_requested = pyqtSignal(str)
    back_requested = pyqtSignal()

    def __init__(self) -> None:
        super().__init__()

        main_layout = QVBoxLayout(self)
        main_layout.setAlignment(Qt.AlignCenter)
        main_layout.setContentsMargins(40, 60, 40, 60)
        main_layout.setSpacing(32)

        # Header
        title = QLabel("Reset Password")
        title.setProperty("heading", True)
        main_layout.addWidget(title, alignment=Qt.AlignCenter)

        subtitle = QLabel("Enter your email to receive reset instructions")
        subtitle.setProperty("caption", True)
        main_layout.addWidget(subtitle, alignment=Qt.AlignCenter)

        # Form Card
        form_card = QFrame()
        form_card.setProperty("card", True)
        form_card.setMaximumWidth(480)
        form_layout = QVBoxLayout(form_card)
        form_layout.setSpacing(20)
        form_layout.setContentsMargins(32, 32, 32, 32)

        email_label = QLabel("Email Address")
        email_label.setStyleSheet(f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;")
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("your.email@company.com")
        form_layout.addWidget(email_label)
        form_layout.addWidget(self.email_input)

        reset_button = QPushButton("Send Reset Instructions")
        reset_button.setProperty("primary", True)
        reset_button.clicked.connect(self._emit_reset_request)
        form_layout.addWidget(reset_button)

        back_button = QPushButton("Back to Sign In")
        back_button.setProperty("flat", True)
        back_button.clicked.connect(self.back_requested)
        form_layout.addWidget(back_button, alignment=Qt.AlignCenter)

        main_layout.addWidget(form_card, alignment=Qt.AlignCenter)
        main_layout.addStretch()

    def _emit_reset_request(self) -> None:
        email = self.email_input.text().strip().lower()
        self.reset_requested.emit(email)


class DashboardPage(QWidget):
    """Modern dashboard with test type organization and file selection."""

    logout_requested = pyqtSignal()
    upload_requested = pyqtSignal(list, str, str)  # file_paths, pump_series, test_type
    refresh_requested = pyqtSignal()
    pump_series_created = pyqtSignal(str, str)
    test_type_created = pyqtSignal(str, str, str)
    files_deleted = pyqtSignal(list)  # List of file IDs to delete
    files_moved = pyqtSignal(list, str, str)  # List of file IDs, new pump series, new test type
    selection_changed = pyqtSignal()  # Signal when pump series or test type changes
    back_to_gateway_requested = pyqtSignal()  # Signal to return to app selection

    def __init__(self) -> None:
        super().__init__()

        # Add pagination variables
        self.current_page = 0
        self.page_size = 50
        self.total_records = 0
        self.all_file_records = []
        self.checkboxes = []  # Track all checkboxes

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(40, 32, 40, 32)
        layout.setSpacing(24)

        # Header Section
        header_layout = QHBoxLayout()
        header_widget = QWidget()
        header_inner = QVBoxLayout(header_widget)
        header_inner.setSpacing(4)
        header_inner.setContentsMargins(0, 0, 0, 0)

        self.welcome_label = QLabel("Dashboard")
        self.welcome_label.setProperty("heading", True)
        header_inner.addWidget(self.welcome_label)

        self.subtitle_label = QLabel("Manage your industrial test data")
        self.subtitle_label.setProperty("caption", True)
        header_inner.addWidget(self.subtitle_label)

        header_layout.addWidget(header_widget)
        header_layout.addStretch()

        # Action Buttons
        action_layout = QHBoxLayout()
        action_layout.setSpacing(12)

        back_button = QPushButton("← Back to App Selection")
        back_button.setProperty("flat", True)
        back_button.clicked.connect(self.back_to_gateway_requested.emit)
        action_layout.addWidget(back_button)

        upload_button = QPushButton("Upload File")
        upload_button.setProperty("primary", True)
        upload_button.clicked.connect(self._select_file)
        action_layout.addWidget(upload_button)

        refresh_button = QPushButton("Refresh")
        refresh_button.setProperty("secondary", True)
        refresh_button.clicked.connect(self.refresh_requested.emit)
        action_layout.addWidget(refresh_button)

        header_layout.addLayout(action_layout)
        layout.addLayout(header_layout)

        # Upload Section Card
        upload_card = QFrame()
        upload_card.setProperty("card", True)
        upload_layout = QVBoxLayout(upload_card)
        upload_layout.setSpacing(16)

        upload_title = QLabel("Pump Series & Test Type")
        upload_title.setProperty("subheading", True)
        upload_layout.addWidget(upload_title)

        # Pump series selector
        pump_series_layout = QHBoxLayout()
        pump_series_layout.setSpacing(12)

        pump_series_label = QLabel("Pump Series:")
        pump_series_label.setStyleSheet(
            f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;"
        )
        pump_series_layout.addWidget(pump_series_label)

        self.pump_series_combo = QComboBox()
        self.pump_series_combo.setMinimumWidth(250)
        self.pump_series_combo.currentIndexChanged.connect(self._handle_pump_series_changed)
        pump_series_layout.addWidget(self.pump_series_combo, stretch=1)

        new_series_button = QPushButton("+ New Pump Series")
        new_series_button.setProperty("secondary", True)
        new_series_button.clicked.connect(self._create_new_pump_series)
        pump_series_layout.addWidget(new_series_button)

        upload_layout.addLayout(pump_series_layout)

        # Test type selector
        test_type_layout = QHBoxLayout()
        test_type_layout.setSpacing(12)

        test_type_label = QLabel("Test Type:")
        test_type_label.setStyleSheet(f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;")
        test_type_layout.addWidget(test_type_label)

        self.test_type_combo = QComboBox()
        self.test_type_combo.setMinimumWidth(250)
        self.test_type_combo.currentIndexChanged.connect(self._handle_test_type_changed)
        test_type_layout.addWidget(self.test_type_combo, stretch=1)

        new_type_button = QPushButton("+ New Test Type")
        new_type_button.setProperty("secondary", True)
        new_type_button.clicked.connect(self._create_new_test_type)
        test_type_layout.addWidget(new_type_button)

        upload_layout.addLayout(test_type_layout)

        info_label = QLabel(
            "Select a test type before uploading files. Files are saved to the shared drive in folders by test type."
        )
        info_label.setProperty("caption", True)
        info_label.setWordWrap(True)
        upload_layout.addWidget(info_label)

        layout.addWidget(upload_card)

        # Files Card
        files_card = QFrame()
        files_card.setProperty("card", True)
        files_layout = QVBoxLayout(files_card)
        files_layout.setContentsMargins(0, 0, 0, 0)
        files_layout.setSpacing(0)

        # Header with select all checkbox
        files_header_widget = QWidget()
        files_header_layout = QHBoxLayout(files_header_widget)
        files_header_layout.setContentsMargins(24, 16, 24, 16)
        files_header_layout.setSpacing(12)

        self.select_all_checkbox = QCheckBox()
        self.select_all_checkbox.stateChanged.connect(self._handle_select_all)
        files_header_layout.addWidget(self.select_all_checkbox)

        files_header = QLabel("Uploaded Files")
        files_header.setProperty("subheading", True)
        files_header_layout.addWidget(files_header)

        files_header_layout.addStretch()

        self.selection_count_label = QLabel("0 selected")
        self.selection_count_label.setProperty("caption", True)
        files_header_layout.addWidget(self.selection_count_label)

        files_header_widget.setStyleSheet(f"border-bottom: 1px solid {DesktopTheme.BORDER};")
        files_layout.addWidget(files_header_widget)

        # Table with checkbox column (now 6 columns instead of 5)
        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(
            [
                "",  # Checkbox column
                "Filename",
                "Pump Series",
                "Test Type",
                "Path",
                "Uploaded",
            ]
        )
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        header.resizeSection(0, 50)  # Fixed width for checkbox column
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.Stretch)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(50)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setFocusPolicy(Qt.NoFocus)
        files_layout.addWidget(self.table)

        # Action buttons with bulk operations
        actions_layout = QHBoxLayout()
        actions_layout.setContentsMargins(16, 12, 16, 16)
        actions_layout.setSpacing(12)

        # Bulk action buttons (only enabled when items selected)
        self.bulk_delete_button = QPushButton("Delete Selected")
        self.bulk_delete_button.setProperty("danger", True)
        self.bulk_delete_button.clicked.connect(self._bulk_delete_files)
        self.bulk_delete_button.setEnabled(False)
        actions_layout.addWidget(self.bulk_delete_button)

        self.bulk_move_button = QPushButton("Move Selected")
        self.bulk_move_button.setProperty("secondary", True)
        self.bulk_move_button.clicked.connect(self._bulk_move_files)
        self.bulk_move_button.setEnabled(False)
        actions_layout.addWidget(self.bulk_move_button)

        actions_layout.addWidget(QLabel("|"))  # Separator

        # Individual file actions
        self.open_file_button = QPushButton("Open File")
        self.open_file_button.setProperty("secondary", True)
        self.open_file_button.clicked.connect(self._open_selected_file)
        actions_layout.addWidget(self.open_file_button)

        self.open_folder_button = QPushButton("Open in Explorer")
        self.open_folder_button.setProperty("secondary", True)
        self.open_folder_button.clicked.connect(self._open_selected_folder)
        actions_layout.addWidget(self.open_folder_button)

        self.copy_path_button = QPushButton("Copy Path")
        self.copy_path_button.setProperty("secondary", True)
        self.copy_path_button.clicked.connect(self._copy_selected_path)
        actions_layout.addWidget(self.copy_path_button)

        self.show_properties_button = QPushButton("Show Properties")
        self.show_properties_button.setProperty("secondary", True)
        self.show_properties_button.clicked.connect(self._show_selected_properties)
        actions_layout.addWidget(self.show_properties_button)

        actions_layout.addStretch()
        files_layout.addLayout(actions_layout)

        layout.addWidget(files_card)

        # CSV Preview Card
        self.csv_card = QFrame()
        self.csv_card.setProperty("card", True)
        csv_layout = QVBoxLayout(self.csv_card)
        csv_layout.setContentsMargins(0, 0, 0, 0)
        csv_layout.setSpacing(0)

        self.csv_preview_label = QLabel("CSV Preview")
        self.csv_preview_label.setProperty("subheading", True)
        self.csv_preview_label.setStyleSheet(
            f"padding: 20px 24px; border-bottom: 1px solid {DesktopTheme.BORDER};"
        )
        csv_layout.addWidget(self.csv_preview_label)

        self.csv_table = QTableWidget(0, 0)
        self.csv_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.csv_table.verticalHeader().setVisible(False)
        self.csv_table.setAlternatingRowColors(True)
        csv_layout.addWidget(self.csv_table)

        self.csv_card.hide()
        layout.addWidget(self.csv_card)

        scroll.setWidget(container)
        main_layout.addWidget(scroll)

        # Add pagination controls after table
        pagination_layout = QHBoxLayout()

        self.prev_button = QPushButton("← Previous")
        self.prev_button.setProperty("secondary", True)
        self.prev_button.clicked.connect(self._load_previous_page)
        pagination_layout.addWidget(self.prev_button)

        self.page_label = QLabel("Page 1 of 1")
        self.page_label.setAlignment(Qt.AlignCenter)
        pagination_layout.addWidget(self.page_label)

        self.next_button = QPushButton("Next →")
        self.next_button.setProperty("secondary", True)
        self.next_button.clicked.connect(self._load_next_page)
        pagination_layout.addWidget(self.next_button)

        self.catalog: Dict[str, List[str]] = {}
        self.pump_series_options: List[str] = []

    def set_catalog(self, catalog: Dict[str, List[str]]) -> None:
        """Update pump series and test type selections."""
        previous_series = self.get_selected_pump_series()
        self.catalog = {name: sorted(types) for name, types in catalog.items()}
        self.pump_series_options = sorted(self.catalog.keys())
        self.pump_series_combo.blockSignals(True)
        self.pump_series_combo.clear()
        if self.pump_series_options:
            self.pump_series_combo.addItems(self.pump_series_options)
            if previous_series in self.pump_series_options:
                index = self.pump_series_combo.findText(previous_series)
                if index >= 0:
                    self.pump_series_combo.setCurrentIndex(index)
        else:
            self.pump_series_combo.addItem("No pump series available")
        self.pump_series_combo.blockSignals(False)
        self._populate_test_types(self.get_selected_pump_series())

    def _populate_test_types(self, pump_series: Optional[str]) -> None:
        self.test_type_combo.blockSignals(True)
        self.test_type_combo.clear()
        if not pump_series or pump_series not in self.catalog:
            self.test_type_combo.addItem("No test types available")
        else:
            test_types = self.catalog.get(pump_series, [])
            if test_types:
                self.test_type_combo.addItems(test_types)
            else:
                self.test_type_combo.addItem("No test types available")
        self.test_type_combo.blockSignals(False)

    def get_selected_test_type(self) -> Optional[str]:
        """Get the currently selected test type."""
        if not self.catalog:
            return None
        value = self.test_type_combo.currentText()
        if value == "No test types available":
            return None
        return value

    def get_selected_pump_series(self) -> Optional[str]:
        if not self.pump_series_options:
            return None
        value = self.pump_series_combo.currentText()
        if value == "No pump series available":
            return None
        return value

    def _handle_pump_series_changed(self) -> None:
        pump_series = self.get_selected_pump_series()
        self._populate_test_types(pump_series)
        self.selection_changed.emit()

    def _handle_test_type_changed(self) -> None:
        """Handle test type selection change."""
        self.selection_changed.emit()

    def _create_new_test_type(self) -> None:
        """Show dialog to create a new test type."""
        pump_series = self.get_selected_pump_series()
        if not pump_series:
            QMessageBox.warning(
                self,
                "Inline Data System",
                "Please select or create a pump series before adding a test type.",
            )
            return
        dialog = NewTestTypeDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            test_type = dialog.get_test_type()
            description = dialog.get_description()
            if test_type:
                self.test_type_created.emit(pump_series, test_type, description)
                # Add to combo box
                test_types = self.catalog.setdefault(pump_series, [])
                if test_type not in test_types:
                    test_types.append(test_type)
                    test_types.sort()
                    self._populate_test_types(pump_series)
                    index = self.test_type_combo.findText(test_type)
                    if index >= 0:
                        self.test_type_combo.setCurrentIndex(index)
                else:
                    QMessageBox.information(
                        self,
                        "Inline Data System",
                        f"Test type '{test_type}' already exists.",
                    )

    def _create_new_pump_series(self) -> None:
        dialog = NewPumpSeriesDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            series_name = dialog.get_pump_series()
            description = dialog.get_description()
            if series_name:
                if series_name in self.catalog:
                    QMessageBox.information(
                        self,
                        "Inline Data System",
                        f"Pump series '{series_name}' already exists.",
                    )
                    return
                self.pump_series_created.emit(series_name, description)
                self.catalog[series_name] = []
                self.pump_series_options = sorted(self.catalog.keys())
                self.pump_series_combo.blockSignals(True)
                self.pump_series_combo.clear()
                self.pump_series_combo.addItems(self.pump_series_options)
                index = self.pump_series_combo.findText(series_name)
                if index >= 0:
                    self.pump_series_combo.setCurrentIndex(index)
                self.pump_series_combo.blockSignals(False)
                self._populate_test_types(series_name)

    def set_user_identity(self, username: str, email: str) -> None:
        username = username.strip()
        # email = email.strip()
        if username:
            self.welcome_label.setText(f"Welcome")

        else:
            self.welcome_label.setText("Dashboard")
            self.subtitle_label.setText("Manage your industrial data")

    # def update_files(self, files: List[Dict[str, Any]]) -> None:
    #     self.file_records = files
    #     self.table.setRowCount(len(files))
    #     for row, file_record in enumerate(files):
    #         filename_item = QTableWidgetItem(file_record.get("filename", ""))
    #         pump_series_item = QTableWidgetItem(file_record.get("pump_series", ""))
    #         test_type_item = QTableWidgetItem(file_record.get("test_type", ""))
    #         path_item = QTableWidgetItem(file_record.get("absolute_path") or file_record.get("file_path", ""))
    #         created_item = QTableWidgetItem(file_record.get("created_at", ""))
    #
    #         for item in (filename_item, pump_series_item, test_type_item, path_item, created_item):
    #             item.setFlags(item.flags() & ~Qt.ItemIsEditable)
    #
    #         self.table.setItem(row, 0, filename_item)
    #         self.table.setItem(row, 1, pump_series_item)
    #         self.table.setItem(row, 2, test_type_item)
    #         self.table.setItem(row, 3, path_item)
    #         self.table.setItem(row, 4, created_item)

    def update_files(self, files: List[Dict[str, Any]]) -> None:
        """Update with pagination"""
        self.all_file_records = files
        self.total_records = len(files)
        self.current_page = 0

        self._display_current_page()

    def _display_current_page(self):
        """Display only the current page of results"""
        start = self.current_page * self.page_size
        end = start + self.page_size

        page_records = self.all_file_records[start:end]

        # Update table with only current page
        self.table.setRowCount(len(page_records))
        for row, file_record in enumerate(page_records):
            filename_item = QTableWidgetItem(file_record.get("filename", ""))
            pump_series_item = QTableWidgetItem(file_record.get("pump_series", ""))
            test_type_item = QTableWidgetItem(file_record.get("test_type", ""))
            path_item = QTableWidgetItem(
                file_record.get("absolute_path") or file_record.get("file_path", "")
            )
            created_item = QTableWidgetItem(file_record.get("created_at", ""))

            for item in (
                filename_item,
                pump_series_item,
                test_type_item,
                path_item,
                created_item,
            ):
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)

            self.table.setItem(row, 0, filename_item)
            self.table.setItem(row, 1, pump_series_item)
            self.table.setItem(row, 2, test_type_item)
            self.table.setItem(row, 3, path_item)
            self.table.setItem(row, 4, created_item)

        # Update pagination controls
        total_pages = (self.total_records + self.page_size - 1) // self.page_size
        self.page_label.setText(f"Page {self.current_page + 1} of {total_pages}")

        self.prev_button.setEnabled(self.current_page > 0)
        self.next_button.setEnabled(self.current_page < total_pages - 1)

    def _get_selected_record(self) -> Optional[Dict[str, Any]]:
        """Get selected record (adjusted for pagination)"""
        selection = self.table.selectionModel().selectedRows()
        if not selection:
            return None
        row = selection[0].row()

        # Adjust for pagination
        actual_index = self.current_page * self.page_size + row
        if 0 <= actual_index < len(self.all_file_records):
            return self.all_file_records[actual_index]
        return None

    def _open_selected_file(self) -> None:
        record = self._get_selected_record()
        if not record:
            QMessageBox.information(self, "Inline Data System", "Select a file first.")
            return
        path_value = record.get("absolute_path") or record.get("file_path")
        if not path_value:
            QMessageBox.warning(self, "Inline Data System", "No file path available.")
            return
        path = Path(path_value)
        if not path.is_absolute():
            base_path = record.get("base_path")
            if base_path:
                path = Path(base_path) / path
        if not path.exists():
            QMessageBox.warning(self, "Inline Data System", f"File not found: {path}")
            return
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(path)))

    def _open_selected_folder(self) -> None:
        record = self._get_selected_record()
        if not record:
            QMessageBox.information(self, "Inline Data System", "Select a file first.")
            return
        path_value = record.get("absolute_path") or record.get("file_path")
        if not path_value:
            QMessageBox.warning(self, "Inline Data System", "No file path available.")
            return
        path = Path(path_value)
        if not path.is_absolute():
            base_path = record.get("base_path")
            if base_path:
                path = Path(base_path) / path
        if not path.exists():
            QMessageBox.warning(self, "Inline Data System", f"File not found: {path}")
            return
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(path.parent)))

    def _copy_selected_path(self) -> None:
        record = self._get_selected_record()
        if not record:
            QMessageBox.information(self, "Inline Data System", "Select a file first.")
            return
        path_value = record.get("absolute_path") or record.get("file_path")
        if not path_value:
            QMessageBox.warning(self, "Inline Data System", "No file path available.")
            return
        path = Path(path_value)
        if not path.is_absolute():
            base_path = record.get("base_path")
            if base_path:
                path = Path(base_path) / path
        QApplication.clipboard().setText(str(path))
        QMessageBox.information(self, "Inline Data System", "File path copied to clipboard.")

    def _load_next_page(self):
        """Load next page"""
        total_pages = (self.total_records + self.page_size - 1) // self.page_size
        if self.current_page < total_pages - 1:
            self.current_page += 1
            self._display_current_page()

    def _load_previous_page(self):
        """Load previous page"""
        if self.current_page > 0:
            self.current_page -= 1
            self._display_current_page()

    def _show_selected_properties(self) -> None:
        record = self._get_selected_record()
        if not record:
            QMessageBox.information(self, "Inline Data System", "Select a file first.")
            return
        path_value = record.get("absolute_path") or record.get("file_path")
        if not path_value:
            QMessageBox.warning(self, "Inline Data System", "No file path available.")
            return
        path = Path(path_value)
        if not path.is_absolute():
            base_path = record.get("base_path")
            if base_path:
                path = Path(base_path) / path
        if not path.exists():
            QMessageBox.warning(self, "Inline Data System", f"File not found: {path}")
            return
        stat = path.stat()
        size_mb = stat.st_size / (1024 * 1024)
        modified = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        created = datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d %H:%M:%S")
        QMessageBox.information(
            self,
            "Inline Data System",
            (
                f"Location: {path}\n"
                f"Size: {size_mb:.2f} MB\n"
                f"Created: {created}\n"
                f"Modified: {modified}"
            ),
        )

    def _select_file(self) -> None:
        pump_series = self.get_selected_pump_series()
        if not pump_series or pump_series == "No pump series available":
            QMessageBox.warning(
                self,
                "Inline Data System",
                "Please select or create a pump series before uploading files.",
            )
            return
        test_type = self.get_selected_test_type()
        if not test_type or test_type == "No test types available":
            QMessageBox.warning(
                self,
                "Inline Data System",
                "Please select or create a test type before uploading files.",
            )
            return

        file_dialog_filter = "Data Files (*.csv *.xlsx *.xlsm *.xltx *.xltm *.asc);;"
        file_dialog_filter += (
            "CSV Files (*.csv);;ASC Files (*.asc);;Excel Files (*.xlsx *.xlsm *.xltx *.xltm)"
        )

        # Changed from getOpenFileName to getOpenFileNames (plural)
        file_paths, _ = QFileDialog.getOpenFileNames(  # <-- THIS IS THE KEY CHANGE
            self,
            "Select Files to Upload",  # Updated text
            "",
            file_dialog_filter,
        )
        if file_paths:  # This is now a list
            self.upload_requested.emit(
                file_paths, pump_series, test_type
            )  # Send list instead of single path

    def display_csv_preview(self, headers: List[str], rows: List[List[str]]) -> None:
        if not headers:
            self.clear_csv_preview()
            return

        column_count = len(headers)
        self.csv_table.setColumnCount(column_count)
        self.csv_table.setHorizontalHeaderLabels(headers)

        self.csv_table.setRowCount(len(rows))
        for row_index, row_values in enumerate(rows):
            for column_index in range(column_count):
                value = row_values[column_index] if column_index < len(row_values) else ""
                item = QTableWidgetItem(value)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.csv_table.setItem(row_index, column_index, item)

        self.csv_card.show()

    def clear_csv_preview(self) -> None:
        self.csv_table.clear()
        self.csv_table.setRowCount(0)
        self.csv_table.setColumnCount(0)
        self.csv_card.hide()

    def _handle_select_all(self, state):
        """Handle select all checkbox state change"""
        is_checked = state == Qt.Checked
        for checkbox in self.checkboxes:
            if checkbox is not None:
                checkbox.blockSignals(True)
                checkbox.setChecked(is_checked)
                checkbox.blockSignals(False)
        self._update_selection_count()

    def _handle_checkbox_change(self):
        """Handle individual checkbox state change"""
        self._update_selection_count()

        # Update select all checkbox state
        checked_count = sum(1 for cb in self.checkboxes if cb and cb.isChecked())
        total_count = len(self.checkboxes)

        self.select_all_checkbox.blockSignals(True)
        if checked_count == 0:
            self.select_all_checkbox.setCheckState(Qt.Unchecked)
        elif checked_count == total_count:
            self.select_all_checkbox.setCheckState(Qt.Checked)
        else:
            self.select_all_checkbox.setCheckState(Qt.PartiallyChecked)
        self.select_all_checkbox.blockSignals(False)

    def _update_selection_count(self):
        """Update the selection count label and enable/disable bulk action buttons"""
        checked_count = sum(1 for cb in self.checkboxes if cb and cb.isChecked())
        self.selection_count_label.setText(f"{checked_count} selected")

        # Enable/disable bulk action buttons
        has_selection = checked_count > 0
        self.bulk_delete_button.setEnabled(has_selection)
        self.bulk_move_button.setEnabled(has_selection)

    def _get_checked_records(self) -> List[Dict[str, Any]]:
        """Get all records that have their checkbox checked"""
        checked_records = []
        for i, checkbox in enumerate(self.checkboxes):
            if checkbox and checkbox.isChecked():
                actual_index = self.current_page * self.page_size + i
                if 0 <= actual_index < len(self.all_file_records):
                    checked_records.append(self.all_file_records[actual_index])
        return checked_records

    def _bulk_delete_files(self):
        """Delete all selected files"""
        checked_records = self._get_checked_records()
        if not checked_records:
            QMessageBox.information(self, "Inline Data System", "No files selected.")
            return

        # Confirm deletion
        reply = QMessageBox.question(
            self,
            "Confirm Deletion",
            f"Are you sure you want to delete {len(checked_records)} file(s)?\n\n"
            "This action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if reply == QMessageBox.Yes:
            file_ids = [record.get("id") for record in checked_records if record.get("id")]
            self.files_deleted.emit(file_ids)

    def _bulk_move_files(self):
        """Move all selected files to a new pump series/test type"""
        checked_records = self._get_checked_records()
        if not checked_records:
            QMessageBox.information(self, "Inline Data System", "No files selected.")
            return

        # Create dialog for selecting new location
        dialog = BulkMoveDialog(self, self.catalog)
        if dialog.exec_() == QDialog.Accepted:
            new_pump_series = dialog.get_pump_series()
            new_test_type = dialog.get_test_type()

            if new_pump_series and new_test_type:
                file_ids = [record.get("id") for record in checked_records if record.get("id")]
                self.files_moved.emit(file_ids, new_pump_series, new_test_type)

    # ============ MODIFIED METHOD FOR DISPLAYING FILES ============

    def _display_current_page(self):
        """Display only the current page of results with checkboxes"""
        start = self.current_page * self.page_size
        end = start + self.page_size

        page_records = self.all_file_records[start:end]

        # Clear old checkboxes
        self.checkboxes = []

        # Update table with only current page
        self.table.setRowCount(len(page_records))
        for row, file_record in enumerate(page_records):
            # Create checkbox widget
            checkbox = QCheckBox()
            checkbox.setMinimumSize(10, 10)  # Add minimum size
            checkbox.setMaximumSize(10, 10)  # Add maximum size to prevent stretching
            checkbox.stateChanged.connect(self._handle_checkbox_change)
            self.checkboxes.append(checkbox)

            # Center the checkbox in the cell
            checkbox_widget = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_widget)
            checkbox_layout.addWidget(checkbox)
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox_layout.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(row, 0, checkbox_widget)

            # Add file data to remaining columns
            filename_item = QTableWidgetItem(file_record.get("filename", ""))
            pump_series_item = QTableWidgetItem(file_record.get("pump_series", ""))
            test_type_item = QTableWidgetItem(file_record.get("test_type", ""))
            path_item = QTableWidgetItem(
                file_record.get("absolute_path") or file_record.get("file_path", "")
            )
            created_item = QTableWidgetItem(file_record.get("created_at", ""))

            for item in (
                filename_item,
                pump_series_item,
                test_type_item,
                path_item,
                created_item,
            ):
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)

            self.table.setItem(row, 1, filename_item)
            self.table.setItem(row, 2, pump_series_item)
            self.table.setItem(row, 3, test_type_item)
            self.table.setItem(row, 4, path_item)
            self.table.setItem(row, 5, created_item)

        # Update pagination controls
        total_pages = (self.total_records + self.page_size - 1) // self.page_size
        self.page_label.setText(f"Page {self.current_page + 1} of {total_pages}")

        self.prev_button.setEnabled(self.current_page > 0)
        self.next_button.setEnabled(self.current_page < total_pages - 1)

        # Reset selection count
        self._update_selection_count()

    def _get_selected_record(self) -> Optional[Dict[str, Any]]:
        """Get selected record (adjusted for pagination) - now uses first checked item or table selection"""
        # First check if any checkboxes are checked
        for i, checkbox in enumerate(self.checkboxes):
            if checkbox and checkbox.isChecked():
                actual_index = self.current_page * self.page_size + i
                if 0 <= actual_index < len(self.all_file_records):
                    return self.all_file_records[actual_index]

        # Fall back to table row selection
        selection = self.table.selectionModel().selectedRows()
        if not selection:
            return None
        row = selection[0].row()

        # Adjust for pagination
        actual_index = self.current_page * self.page_size + row
        if 0 <= actual_index < len(self.all_file_records):
            return self.all_file_records[actual_index]
        return None


# ============ NEW DIALOG FOR BULK MOVE ============


class BulkMoveDialog(QDialog):
    """Dialog for moving files to a new pump series/test type."""

    def __init__(self, parent=None, catalog=None):
        super().__init__(parent)
        self.catalog = catalog or {}
        self.setWindowTitle("Move Files")
        self.setMinimumWidth(400)

        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(24, 24, 24, 24)

        title = QLabel("Move Selected Files")
        title.setProperty("subheading", True)
        layout.addWidget(title)

        desc = QLabel("Select the destination pump series and test type for the selected files.")
        desc.setProperty("caption", True)
        desc.setWordWrap(True)
        layout.addWidget(desc)

        # Pump series selector
        pump_label = QLabel("Destination Pump Series")
        pump_label.setStyleSheet(f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;")
        layout.addWidget(pump_label)

        self.pump_combo = QComboBox()
        self.pump_combo.addItems(sorted(self.catalog.keys()))
        self.pump_combo.currentIndexChanged.connect(self._update_test_types)
        layout.addWidget(self.pump_combo)

        # Test type selector
        test_label = QLabel("Destination Test Type")
        test_label.setStyleSheet(f"color: {DesktopTheme.TEXT_SECONDARY}; font-weight: 500;")
        layout.addWidget(test_label)

        self.test_combo = QComboBox()
        layout.addWidget(self.test_combo)

        # Initialize test types
        self._update_test_types()

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)

        for button in button_box.buttons():
            if button_box.buttonRole(button) == QDialogButtonBox.AcceptRole:
                button.setProperty("primary", True)
            else:
                button.setProperty("secondary", True)
            button.setMinimumHeight(44)

        layout.addWidget(button_box)

    def _update_test_types(self):
        """Update test type combo based on selected pump series"""
        self.test_combo.clear()
        pump_series = self.pump_combo.currentText()
        if pump_series in self.catalog:
            self.test_combo.addItems(self.catalog[pump_series])

    def get_pump_series(self) -> str:
        return self.pump_combo.currentText().strip()

    def get_test_type(self) -> str:
        return self.test_combo.currentText().strip()


class IndustrialDataApp(QMainWindow):
    """Main window with modern industrial UI design."""

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Inline Data System")
        self.resize(1200, 800)
        self.setMinimumSize(900, 600)
        self.setAttribute(Qt.WA_DeleteOnClose, True)

        self.session_state = SessionState()
        self.db_manager = DatabaseManager()
        self.auth_store = LocalAuthStore(self.db_manager)
        self.history_store = UploadHistoryStore(self.db_manager)
        self.storage_manager = LocalStorageManager(config=CONFIG, database=self.db_manager)
        self.current_username: str = ""
        self.default_pump_series = "General"

        # Create menu bar
        self._create_menu_bar()

        # Create toolbar
        self._create_toolbar()

        # Create status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Ready")

        self.dashboard_page = DashboardPage()
        self.setCentralWidget(self.dashboard_page)
        self.dashboard_page.upload_requested.connect(self.handle_upload)
        self.dashboard_page.refresh_requested.connect(self.refresh_files)
        self.dashboard_page.pump_series_created.connect(self.handle_new_pump_series)
        self.dashboard_page.test_type_created.connect(self.handle_new_test_type)
        self.dashboard_page.files_deleted.connect(self.handle_delete_files)
        self.dashboard_page.selection_changed.connect(self.refresh_files)
        self.dashboard_page.back_to_gateway_requested.connect(self.close)

        self._initialize_gateway_session()

    def _create_menu_bar(self) -> None:
        """Create traditional desktop menu bar."""
        menubar = self.menuBar()

        # File Menu
        file_menu = menubar.addMenu("&File")

        new_pump_action = QAction("New &Pump Series...", self)
        new_pump_action.setShortcut(QKeySequence("Ctrl+P"))
        new_pump_action.triggered.connect(lambda: self.dashboard_page.create_pump_series_clicked.emit())
        file_menu.addAction(new_pump_action)

        new_test_action = QAction("New &Test Type...", self)
        new_test_action.setShortcut(QKeySequence("Ctrl+T"))
        new_test_action.triggered.connect(lambda: self.dashboard_page.create_test_type_clicked.emit())
        file_menu.addAction(new_test_action)

        file_menu.addSeparator()

        upload_action = QAction("&Upload Files...", self)
        upload_action.setShortcut(QKeySequence("Ctrl+U"))
        upload_action.triggered.connect(lambda: self.dashboard_page.upload_clicked.emit())
        file_menu.addAction(upload_action)

        file_menu.addSeparator()

        exit_action = QAction("E&xit", self)
        exit_action.setShortcut(QKeySequence("Alt+F4"))
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # Edit Menu
        edit_menu = menubar.addMenu("&Edit")

        refresh_action = QAction("&Refresh", self)
        refresh_action.setShortcut(QKeySequence("F5"))
        refresh_action.triggered.connect(lambda: self.dashboard_page.refresh_clicked.emit())
        edit_menu.addAction(refresh_action)

        # View Menu
        view_menu = menubar.addMenu("&View")

        show_toolbar_action = QAction("Show &Toolbar", self)
        show_toolbar_action.setCheckable(True)
        show_toolbar_action.setChecked(True)
        show_toolbar_action.triggered.connect(lambda checked: self.toolbar.setVisible(checked))
        view_menu.addAction(show_toolbar_action)

        show_statusbar_action = QAction("Show &Status Bar", self)
        show_statusbar_action.setCheckable(True)
        show_statusbar_action.setChecked(True)
        show_statusbar_action.triggered.connect(lambda checked: self.status_bar.setVisible(checked))
        view_menu.addAction(show_statusbar_action)

        # Help Menu
        help_menu = menubar.addMenu("&Help")

        about_action = QAction("&About", self)
        about_action.triggered.connect(self._show_about_dialog)
        help_menu.addAction(about_action)

    def _create_toolbar(self) -> None:
        """Create traditional desktop toolbar."""
        self.toolbar = QToolBar("Main Toolbar")
        self.toolbar.setMovable(False)
        self.addToolBar(self.toolbar)

        # Upload button
        upload_btn = QAction("Upload", self)
        upload_btn.setToolTip("Upload Files (Ctrl+U)")
        upload_btn.triggered.connect(lambda: self.dashboard_page.upload_clicked.emit())
        self.toolbar.addAction(upload_btn)

        self.toolbar.addSeparator()

        # Refresh button
        refresh_btn = QAction("Refresh", self)
        refresh_btn.setToolTip("Refresh File List (F5)")
        refresh_btn.triggered.connect(lambda: self.dashboard_page.refresh_clicked.emit())
        self.toolbar.addAction(refresh_btn)

        self.toolbar.addSeparator()

        # New Pump Series button
        new_pump_btn = QAction("New Pump", self)
        new_pump_btn.setToolTip("Create New Pump Series (Ctrl+P)")
        new_pump_btn.triggered.connect(lambda: self.dashboard_page.create_pump_series_clicked.emit())
        self.toolbar.addAction(new_pump_btn)

        # New Test Type button
        new_test_btn = QAction("New Test", self)
        new_test_btn.setToolTip("Create New Test Type (Ctrl+T)")
        new_test_btn.triggered.connect(lambda: self.dashboard_page.create_test_type_clicked.emit())
        self.toolbar.addAction(new_test_btn)

    def _show_about_dialog(self) -> None:
        """Show about dialog."""
        QMessageBox.about(
            self,
            "About Inline Data System",
            "<h3>Inline Data System</h3>"
            "<p>Version 1.0</p>"
            "<p>Industrial data management and analysis system.</p>"
        )

    def _initialize_gateway_session(self) -> None:
        gateway_user = self._ensure_gateway_user()
        self._set_logged_in_user(gateway_user)

    def _ensure_gateway_user(self) -> LocalUser:
        """Create or retrieve the shared gateway account for upload access."""

        default_email = os.getenv("IDS_GATEWAY_USER_EMAIL") or "gateway@local"
        default_username = os.getenv("IDS_GATEWAY_USERNAME") or "gateway"
        default_password = os.getenv("IDS_GATEWAY_PASSWORD") or "gateway"
        default_display_name = os.getenv("IDS_GATEWAY_DISPLAY_NAME") or "Gateway Access"

        record = self.db_manager.get_user_by_email(default_email)
        if record is None and default_username:
            record = self.db_manager.get_user_by_username(default_username)

        if record is not None:
            metadata = dict(record.metadata or {})
            if default_username and "username" not in metadata and record.username:
                metadata["username"] = record.username
            if default_username and "username_normalized" not in metadata and record.username:
                metadata["username_normalized"] = record.username.lower()
            if "display_name" not in metadata:
                metadata["display_name"] = default_display_name
            if metadata != record.metadata:
                self.db_manager.update_user(record.id, metadata=metadata)
                refreshed = self.db_manager.get_user_by_id(record.id)
                if refreshed is not None:
                    record = refreshed
            return LocalUser(
                id=record.id,
                email=record.email,
                username=record.username,
                password_hash=record.password_hash,
                salt=record.salt,
                metadata=record.metadata,
                created_at=record.created_at,
            )

        metadata = {
            "display_name": default_display_name,
            "username": default_username,
            "username_normalized": (
                default_username.lower() if default_username else default_username
            ),
        }

        try:
            return self.auth_store.create_user(
                email=default_email,
                password=default_password,
                username=default_username,
                metadata=metadata,
            )
        except ValueError:
            fallback = self.db_manager.get_user_by_email(default_email)
            if fallback is None and default_username:
                fallback = self.db_manager.get_user_by_username(default_username)
            if fallback is None:
                raise
            return LocalUser(
                id=fallback.id,
                email=fallback.email,
                username=fallback.username,
                password_hash=fallback.password_hash,
                salt=fallback.salt,
                metadata=fallback.metadata,
                created_at=fallback.created_at,
            )

    def load_test_types(self) -> None:
        """Load available test types from the database and shared drive."""
        try:
            catalog: Dict[str, set[str]] = {}

            def ensure_series(name: Optional[str]) -> set[str]:
                series_name = (name or self.default_pump_series).strip() or self.default_pump_series
                return catalog.setdefault(series_name, set())

            # Existing pump series from database
            for series_record in self.db_manager.list_pump_series():
                ensure_series(series_record.name)

            # Existing test types from database
            for record in self.db_manager.list_test_types():
                ensure_series(record.pump_series).add(record.name)

            # Scan filesystem for additional pump series/test types
            base_dir = CONFIG.files_base_path
            if base_dir.exists():
                legacy_tests_dir = base_dir / "tests"
                if legacy_tests_dir.exists():
                    legacy_bucket = ensure_series(self.default_pump_series)
                    for child in legacy_tests_dir.iterdir():
                        if child.is_dir():
                            legacy_bucket.add(child.name)
                for series_dir in base_dir.iterdir():
                    if not series_dir.is_dir():
                        continue
                    if series_dir.name == "tests":
                        continue
                    series_bucket = ensure_series(series_dir.name)
                    tests_dir = series_dir / "tests"
                    if tests_dir.exists():
                        for child in tests_dir.iterdir():
                            if child.is_dir():
                                series_bucket.add(child.name)

            if not catalog:
                ensure_series(self.default_pump_series)

            normalized_catalog = {name: sorted(types) for name, types in catalog.items()}
            self.dashboard_page.set_catalog(normalized_catalog)
        except Exception:
            self.dashboard_page.set_catalog({})

    def handle_new_pump_series(self, name: str, description: str) -> None:
        name = name.strip()
        if not name:
            self._alert("Pump series name is required.", QMessageBox.Warning)
            return
        description_value = description.strip() or None
        try:
            record = self.db_manager.ensure_pump_series(name, description_value)
            self.storage_manager.ensure_pump_series_exists(record.name)
        except StorageError as exc:
            self._alert(str(exc), QMessageBox.Warning)
            return
        except Exception as exc:
            self._alert(f"Unable to create pump series: {exc}", QMessageBox.Critical)
            return
        self.load_test_types()
        index = self.dashboard_page.pump_series_combo.findText(record.name)
        if index >= 0:
            self.dashboard_page.pump_series_combo.setCurrentIndex(index)

    def handle_new_test_type(self, pump_series: str, name: str, description: str) -> None:
        pump_series = pump_series.strip()
        name = name.strip()
        if not pump_series:
            pump_series = self.default_pump_series
        description_value = description.strip() or None
        try:
            record = self.db_manager.ensure_test_type(
                name, description_value, pump_series=pump_series
            )
            self.storage_manager.ensure_folder_exists(pump_series, record.name)
        except StorageError as exc:
            self._alert(str(exc), QMessageBox.Warning)
            return
        except Exception as exc:
            self._alert(f"Unable to create test type: {exc}", QMessageBox.Critical)
            return
        self.load_test_types()
        series_index = self.dashboard_page.pump_series_combo.findText(pump_series)
        if series_index >= 0:
            self.dashboard_page.pump_series_combo.setCurrentIndex(series_index)
        index = self.dashboard_page.test_type_combo.findText(record.name)
        if index >= 0:
            self.dashboard_page.test_type_combo.setCurrentIndex(index)

    def handle_delete_files(self, file_ids: List[int]) -> None:
        """Handle deletion of files from storage and database."""
        if not file_ids:
            return

        deleted_count = 0
        failed_count = 0

        for file_id in file_ids:
            try:
                # Get the file record from database to get the file path
                upload_record = self.db_manager.get_upload_by_id(file_id)
                if upload_record:
                    # Delete from storage if path exists
                    file_path = upload_record.get("file_path")
                    if file_path:
                        absolute_path = CONFIG.files_base_path / Path(file_path)
                        if absolute_path.exists():
                            try:
                                self.storage_manager.delete_file(absolute_path)
                            except StorageError:
                                # Continue even if file doesn't exist or can't be deleted
                                pass

                # Delete from database
                self.db_manager.delete_upload(file_id)
                deleted_count += 1
            except Exception:
                failed_count += 1

        # Show result message
        if deleted_count > 0:
            message = f"Successfully deleted {deleted_count} file(s)."
            if failed_count > 0:
                message += f"\nFailed to delete {failed_count} file(s)."
            self._alert(message, QMessageBox.Information)
        elif failed_count > 0:
            self._alert(f"Failed to delete {failed_count} file(s).", QMessageBox.Warning)

        # Refresh the file list
        self.refresh_files()

    def _set_logged_in_user(self, user: LocalUser) -> None:
        session_payload = {
            "id": user.id,
            "email": user.email,
            "username": (
                user.username
                or user.metadata.get("username")
                or user.metadata.get("username_normalized")
                or ""
            ),
            "metadata": user.metadata,
        }
        self.session_state.set_user(session_payload)
        self.current_username = session_payload["username"] or ""
        self.dashboard_page.set_user_identity(
            self.current_username, session_payload.get("email", "")
        )
        self.refresh_files()

    def refresh_files(self) -> None:
        user = self.session_state.user
        if not user:
            self._alert(
                "Gateway session was reset. Restoring access...",
                QMessageBox.Information,
            )
            self._initialize_gateway_session()
            return

        metadata = user.get("metadata") or {}
        username = (
            user.get("username")
            or metadata.get("username")
            or metadata.get("username_normalized")
            or self.current_username
            or ""
        ).strip()
        self.current_username = username

        self.dashboard_page.set_user_identity(
            self.current_username,
            user.get("email", ""),
        )

        user_id = user.get("id")
        if user_id is None:
            self._alert("Session error: missing user identifier.", QMessageBox.Warning)
            self._initialize_gateway_session()
            return

        base_path = self.storage_manager.base_path
        if base_path.exists():
            self.db_manager.prune_missing_uploads(base_path)

        # Get selected pump series and test type for filtering
        selected_pump_series = self.dashboard_page.get_selected_pump_series()
        selected_test_type = self.dashboard_page.get_selected_test_type()

        records = []
        for record in self.history_store.get_records_for_user(int(user_id)):
            relative_path = record.get("file_path")
            absolute_path = None
            if relative_path:
                absolute_candidate = (CONFIG.files_base_path / Path(relative_path)).resolve()
                absolute_path = str(absolute_candidate)
            record["absolute_path"] = absolute_path
            record["base_path"] = str(CONFIG.files_base_path)
            record["pump_series"] = record.get("pump_series") or self.default_pump_series

            # Filter by selected pump series and test type
            record_pump_series = record.get("pump_series")
            record_test_type = record.get("test_type")

            # Only include records that match the selected filters
            if selected_pump_series and record_pump_series != selected_pump_series:
                continue
            if selected_test_type and record_test_type != selected_test_type:
                continue

            records.append(record)

        self.dashboard_page.update_files(records)
        self.load_test_types()

    def handle_upload(
        self,
        file_paths: str | List[str],
        pump_series: str,
        test_type: str,
    ) -> None:
        # Convert single file to list for uniform handling
        if isinstance(file_paths, str):
            file_paths = [file_paths]

        user = self.session_state.user
        if not user:
            self._alert(
                "Gateway session was reset. Restoring access...",
                QMessageBox.Information,
            )
            self._initialize_gateway_session()
            return

        pump_series = pump_series.strip()
        if not pump_series:
            self._alert("Please select a pump series.", QMessageBox.Warning)
            return

        test_type = test_type.strip()
        if not test_type:
            self._alert("Please select a test type.", QMessageBox.Warning)
            return

        supported_extensions = {".csv", ".xlsx", ".xlsm", ".xltx", ".xltm", ".asc"}

        successful_uploads = []
        failed_uploads = []

        for file_path in file_paths:
            file_extension = Path(file_path).suffix.lower()

            if file_extension not in supported_extensions:
                allowed = ", ".join(sorted(supported_extensions))
                failed_uploads.append((file_path, f"Unsupported file type. Allowed: {allowed}"))
                continue

            # Only show preview for single file
            if len(file_paths) == 1:
                preview_result = self._prepare_file_preview(file_path, file_extension)
                if preview_result is None:
                    failed_uploads.append((file_path, "Preview generation failed"))
                    continue
                headers, rows = preview_result
                self.dashboard_page.display_csv_preview(headers, rows)
            else:
                self.dashboard_page.clear_csv_preview()

            try:
                stored = self.storage_manager.upload_file(file_path, pump_series, test_type)
            except StorageError as exc:
                failed_uploads.append((file_path, str(exc)))
                continue

            try:
                self.history_store.add_record(
                    user_id=int(user.get("id")),
                    filename=os.path.basename(file_path),
                    file_path=str(stored.relative_path),
                    pump_series=pump_series,
                    test_type=test_type,
                    file_size=stored.size_bytes,
                )
                successful_uploads.append(file_path)
            except Exception as exc:
                # Attempt to clean up the copied file if database write fails
                try:
                    self.storage_manager.delete_file(stored.absolute_path)
                except StorageError:
                    pass
                failed_uploads.append((file_path, f"Failed to record upload: {exc}"))
                continue

        # Show summary message
        if len(file_paths) > 1:
            summary = f"Upload Complete:\n"
            summary += f"✓ Successfully uploaded: {len(successful_uploads)} files\n"
            if failed_uploads:
                summary += f"✗ Failed: {len(failed_uploads)} files\n\n"
                summary += "Failed files:\n"
                for fname, error in failed_uploads[:5]:  # Show first 5 failures
                    summary += f"  - {Path(fname).name}: {error}\n"
                if len(failed_uploads) > 5:
                    summary += f"  ... and {len(failed_uploads) - 5} more\n"
            self._alert(
                summary,
                QMessageBox.Information if successful_uploads else QMessageBox.Warning,
            )
        elif successful_uploads:
            # Single file success message
            stored_path = self.storage_manager.get_file_path(
                pump_series,
                test_type,
                os.path.basename(successful_uploads[0]),
            )
            message = f"File uploaded to shared drive at:\n{stored_path}"

            # Add note about parquet conversion for ASC files
            if Path(successful_uploads[0]).suffix.lower() == ".asc":
                parquet_path = stored_path.with_suffix(".parquet")
                message += f"\n\n✓ ASC file converted to Parquet format:\n{parquet_path}"

            self._alert(message, QMessageBox.Information)
        elif failed_uploads:
            # Single file failure
            self._alert(failed_uploads[0][1], QMessageBox.Critical)

        if successful_uploads:
            self.refresh_files()

    @staticmethod
    def _is_valid_email(email: str) -> bool:
        return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email))

    def _prepare_file_preview(
        self, file_path: str, file_extension: str
    ) -> Optional[tuple[List[str], List[List[str]]]]:
        rows: List[List[str]] = []
        if file_extension == ".csv" or file_extension == ".asc":
            try:
                # Try UTF-8 first, fall back to latin-1 or cp1252
                encodings = ["utf-8", "latin-1", "cp1252", "iso-8859-1"]
                raw_text = ""
                for encoding in encodings:
                    try:
                        with open(file_path, encoding=encoding) as file:
                            raw_text = file.read()
                        break  # Success, exit loop
                    except UnicodeDecodeError:
                        if encoding == encodings[-1]:
                            raise
                        continue  # Try next encoding

                if not raw_text.strip():
                    rows = []
                else:
                    # ``ASC`` exports frequently contain leading metadata lines.
                    # Keep only rows that look tabular before running the CSV
                    # sniffer so the preview focuses on the structured data.
                    candidate_lines: List[str] = []
                    for line in raw_text.splitlines():
                        if not line.strip():
                            continue
                        # Identify common delimiters
                        if any(delimiter in line for delimiter in ("\t", ";", ",", "|")):
                            candidate_lines.append(line)

                    sample_text = "\n".join(candidate_lines[:40]) or raw_text[:4096]

                    try:
                        dialect = csv.Sniffer().sniff(sample_text, delimiters=[",", ";", "\t", "|"])
                    except csv.Error:
                        dialect = csv.excel_tab if file_extension == ".asc" else csv.excel

                    reader = csv.reader(
                        io.StringIO("\n".join(candidate_lines) if candidate_lines else raw_text),
                        dialect,
                    )
                    rows = [row for row in reader if any(cell.strip() for cell in row)]
            except Exception as exc:
                self._alert(f"Unable to read file: {exc}", QMessageBox.Critical)
                self.dashboard_page.clear_csv_preview()
                return None
        else:

            try:
                from openpyxl import load_workbook
            except ImportError:
                self._alert(
                    "Excel support is unavailable because openpyxl is not installed.",
                    QMessageBox.Critical,
                )
                self.dashboard_page.clear_csv_preview()
                return None
            workbook = None
            try:
                workbook = load_workbook(
                    filename=file_path,
                    read_only=True,
                    data_only=True,
                )
                worksheet = workbook.active
                for row in worksheet.iter_rows(values_only=True):
                    rows.append(["" if cell is None else str(cell) for cell in row])
                    if len(rows) >= MAX_PREVIEW_ROWS + 1:
                        rows = rows[: MAX_PREVIEW_ROWS + 1]
                        break
            except Exception as exc:
                self._alert(
                    f"Unable to read Excel file: {exc}",
                    QMessageBox.Critical,
                )
                self.dashboard_page.clear_csv_preview()
                return None
            finally:
                if workbook is not None:
                    try:
                        workbook.close()
                    except Exception:
                        pass

        if not rows:
            self._alert("The selected file is empty.", QMessageBox.Warning)
            self.dashboard_page.clear_csv_preview()
            return None

        headers = [str(value) for value in rows[0]] if rows else []
        data_rows = [[str(value) for value in row] for row in rows[1:101]]

        # Remove empty columns - keep only columns that have data
        if headers and data_rows:
            num_cols = len(headers)
            non_empty_col_indices = []

            for col_idx in range(num_cols):
                # Check if header is non-empty
                header_has_content = headers[col_idx].strip() != ""

                # Check if any data cell in this column has content
                col_has_data = any(
                    row[col_idx].strip() != "" for row in data_rows if col_idx < len(row)
                )

                # Keep column if header or any data cell has content
                if header_has_content or col_has_data:
                    non_empty_col_indices.append(col_idx)

            # Filter headers and data rows to keep only non-empty columns
            if non_empty_col_indices:
                headers = [headers[i] for i in non_empty_col_indices]
                data_rows = [
                    [row[i] if i < len(row) else "" for i in non_empty_col_indices]
                    for row in data_rows
                ]

        return headers, data_rows

    def _alert(self, message: str, icon: QMessageBox.Icon) -> None:
        dialog = QMessageBox(self)
        dialog.setIcon(icon)
        dialog.setText(message)
        dialog.setWindowTitle("Inline Data System")
        dialog.setStyleSheet(
            f"""
            QMessageBox {{
                background-color: {DesktopTheme.SURFACE};
            }}
            QMessageBox QLabel {{
                color: {DesktopTheme.TEXT_PRIMARY};
                font-size: 14px;
                padding: 10px;
            }}
            QPushButton {{
                min-width: 80px;
                padding: 8px 16px;
            }}
        """
        )
        dialog.exec_()


def main() -> None:
    app = QApplication(sys.argv)

    # Apply industrial theme stylesheet
    app.setStyleSheet(DesktopTheme.get_stylesheet())

    # Set application-wide font
    app.setFont(QFont("Segoe UI", 10))

    window = IndustrialDataApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
