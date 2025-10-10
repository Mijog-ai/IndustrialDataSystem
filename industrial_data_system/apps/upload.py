"""Enhanced PyQt5 UI with test type organization for uploads."""
from __future__ import annotations

import csv
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from dotenv import load_dotenv
from PyQt5.QtCore import Qt, pyqtSignal, QSize, QUrl
from PyQt5.QtGui import QDesktopServices, QFont, QPalette, QColor, QIcon
from PyQt5.QtWidgets import (
    QApplication,
    QFileDialog,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QFrame,
    QScrollArea,
    QComboBox,
    QDialog,
    QDialogButtonBox,
)
from industrial_data_system.core.auth import LocalAuthStore, LocalUser, UploadHistoryStore
from industrial_data_system.core.config import get_config
from industrial_data_system.core.db_manager import DatabaseManager
from industrial_data_system.core.storage import LocalStorageManager, StorageError

# ---------------------------------------------------------------------------
# Environment loading helpers
# ---------------------------------------------------------------------------


def _load_environment() -> None:
    """Load environment variables with support for frozen executables."""

    candidate_paths = []

    # 1. Directory that contains the running script (useful in development).
    script_directory = Path(__file__).resolve().parent
    candidate_paths.append(script_directory / ".env")
    candidate_paths.append(script_directory.parent / ".env")
    candidate_paths.append(script_directory.parent.parent / ".env")

    # 2. When running from a PyInstaller bundle, assets live under ``_MEIPASS``.
    meipass_dir = getattr(sys, "_MEIPASS", None)
    if meipass_dir:
        candidate_paths.append(Path(meipass_dir) / ".env")

    # 3. Current working directory (covers the case where the .exe is launched
    #    from a shortcut or another folder but the .env sits next to it).
    candidate_paths.append(Path.cwd() / ".env")

    for env_path in candidate_paths:
        if env_path.is_file():
            load_dotenv(env_path)
            break
    else:
        # Fall back to the default search behaviour which looks up the tree.
        load_dotenv()


_load_environment()

CONFIG = get_config()


class IndustrialTheme:
    """Industrial design system color palette and styles."""

    # Color Palette
    PRIMARY = "#1E3A8A"  # Deep Blue
    PRIMARY_LIGHT = "#3B82F6"  # Light Blue
    PRIMARY_DARK = "#1E40AF"  # Darker Blue

    SECONDARY = "#64748B"  # Slate Gray
    SECONDARY_LIGHT = "#94A3B8"

    SUCCESS = "#10B981"  # Green
    WARNING = "#F59E0B"  # Amber
    ERROR = "#EF4444"  # Red

    BACKGROUND = "#F8FAFC"  # Light Gray
    SURFACE = "#FFFFFF"  # White
    SURFACE_DARK = "#F1F5F9"

    TEXT_PRIMARY = "#0F172A"  # Dark Slate
    TEXT_SECONDARY = "#475569"  # Medium Slate
    TEXT_HINT = "#94A3B8"  # Light Slate

    BORDER = "#E2E8F0"
    BORDER_FOCUS = "#3B82F6"

    @staticmethod
    def get_stylesheet():
        """Return complete application stylesheet."""
        return f"""
            QMainWindow {{
                background-color: {IndustrialTheme.BACKGROUND};
            }}
            
            QWidget {{
                font-family: 'Segoe UI', 'San Francisco', 'Helvetica Neue', Arial, sans-serif;
                font-size: 14px;
                color: {IndustrialTheme.TEXT_PRIMARY};
            }}
            
            QLabel {{
                color: {IndustrialTheme.TEXT_PRIMARY};
            }}
            
            QLabel[heading="true"] {{
                font-size: 28px;
                font-weight: 600;
                color: {IndustrialTheme.TEXT_PRIMARY};
                padding: 8px 0px;
            }}
            
            QLabel[subheading="true"] {{
                font-size: 18px;
                font-weight: 500;
                color: {IndustrialTheme.TEXT_PRIMARY};
                padding: 4px 0px;
            }}
            
            QLabel[caption="true"] {{
                font-size: 12px;
                color: {IndustrialTheme.TEXT_SECONDARY};
            }}
            
            QLineEdit {{
                padding: 12px 16px;
                border: 2px solid {IndustrialTheme.BORDER};
                border-radius: 8px;
                background-color: {IndustrialTheme.SURFACE};
                color: {IndustrialTheme.TEXT_PRIMARY};
                font-size: 14px;
            }}
            
            
            
            QLineEdit:disabled {{
                background-color: {IndustrialTheme.SURFACE_DARK};
                color: {IndustrialTheme.TEXT_HINT};
            }}
            
            QComboBox {{
                padding: 12px 16px;
                border: 2px solid {IndustrialTheme.BORDER};
                border-radius: 8px;
                background-color: {IndustrialTheme.SURFACE};
                color: {IndustrialTheme.TEXT_PRIMARY};
                font-size: 14px;
                min-height: 44px;
            }}
            
           
            
            QComboBox::drop-down {{
                border: none;
                width: 30px;
            }}
            
            QComboBox QAbstractItemView {{
                background-color: {IndustrialTheme.SURFACE};
                border: 2px solid {IndustrialTheme.BORDER};
                border-radius: 8px;
                selection-background-color: {IndustrialTheme.PRIMARY_LIGHT};
                padding: 4px;
            }}
            
            QPushButton {{
                padding: 12px 24px;
                border: none;
                border-radius: 8px;
                font-size: 14px;
                font-weight: 500;
                min-height: 44px;
            }}
            
            QPushButton[primary="true"] {{
                background-color: {IndustrialTheme.PRIMARY};
                color: white;
            }}
            
            QPushButton[primary="true"]:hover {{
                background-color: {IndustrialTheme.PRIMARY_DARK};
            }}
            
            QPushButton[primary="true"]:pressed {{
                background-color: {IndustrialTheme.PRIMARY_DARK};
            }}
            
            QPushButton[secondary="true"] {{
                background-color: {IndustrialTheme.SURFACE};
                color: {IndustrialTheme.TEXT_PRIMARY};
                border: 2px solid {IndustrialTheme.BORDER};
            }}
            
            QPushButton[secondary="true"]:hover {{
                background-color: {IndustrialTheme.SURFACE_DARK};
                border: 2px solid {IndustrialTheme.SECONDARY};
            }}
            
            QPushButton[danger="true"] {{
                background-color: {IndustrialTheme.ERROR};
                color: white;
            }}
            
            QPushButton[danger="true"]:hover {{
                background-color: #DC2626;
            }}
            
            QPushButton[flat="true"] {{
                background-color: transparent;
                color: {IndustrialTheme.PRIMARY};
                padding: 8px 16px;
            }}
            
            QPushButton[flat="true"]:hover {{
                background-color: rgba(59, 130, 246, 0.1);
            }}
            
            QPushButton:disabled {{
                background-color: {IndustrialTheme.SURFACE_DARK};
                color: {IndustrialTheme.TEXT_HINT};
            }}
            
            QTableWidget {{
                background-color: {IndustrialTheme.SURFACE};
                border: 1px solid {IndustrialTheme.BORDER};
                border-radius: 8px;
                gridline-color: {IndustrialTheme.BORDER};
            }}
            
            QTableWidget::item {{
                padding: 12px;
                border-bottom: 1px solid {IndustrialTheme.BORDER};
            }}
            
            QTableWidget::item:selected {{
                background-color: rgba(59, 130, 246, 0.1);
                color: {IndustrialTheme.TEXT_PRIMARY};
            }}
            
            QHeaderView::section {{
                background-color: {IndustrialTheme.SURFACE_DARK};
                padding: 12px;
                border: none;
                border-bottom: 2px solid {IndustrialTheme.BORDER};
                font-weight: 600;
                color: {IndustrialTheme.TEXT_PRIMARY};
            }}
            
            QFrame[card="true"] {{
                background-color: {IndustrialTheme.SURFACE};
                border: 1px solid {IndustrialTheme.BORDER};
                border-radius: 12px;
                padding: 24px;
            }}
            
            QDialog {{
                background-color: {IndustrialTheme.SURFACE};
            }}
            
            QScrollBar:vertical {{
                border: none;
                background-color: {IndustrialTheme.SURFACE_DARK};
                width: 12px;
                border-radius: 6px;
            }}
            
            QScrollBar::handle:vertical {{
                background-color: {IndustrialTheme.SECONDARY_LIGHT};
                border-radius: 6px;
                min-height: 30px;
            }}
            
            QScrollBar::handle:vertical:hover {{
                background-color: {IndustrialTheme.SECONDARY};
            }}
            
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
        """


class NewTestTypeDialog(QDialog):
    """Dialog for creating a new test type."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Create New Test Type")
        self.setMinimumWidth(400)

        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(24, 24, 24, 24)

        # Title
        title = QLabel("New Test Type")
        title.setProperty("subheading", True)
        layout.addWidget(title)

        # Description
        desc = QLabel(
            "Enter a name for the new test type. A matching folder will be created on the shared drive."
        )
        desc.setProperty("caption", True)
        desc.setWordWrap(True)
        layout.addWidget(desc)

        # Input field
        input_label = QLabel("Test Type Name")
        input_label.setStyleSheet(f"color: {IndustrialTheme.TEXT_SECONDARY}; font-weight: 500;")
        layout.addWidget(input_label)

        self.test_type_input = QLineEdit()
        self.test_type_input.setPlaceholderText("e.g., Performance Test, Load Test, Stress Test")
        layout.addWidget(self.test_type_input)

        description_label = QLabel("Description (optional)")
        description_label.setStyleSheet(
            f"color: {IndustrialTheme.TEXT_SECONDARY}; font-weight: 500;"
        )
        layout.addWidget(description_label)

        self.description_input = QLineEdit()
        self.description_input.setPlaceholderText("Short summary for this test type")
        layout.addWidget(self.description_input)

        # Buttons
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)

        # Style buttons
        for button in button_box.buttons():
            if button_box.buttonRole(button) == QDialogButtonBox.AcceptRole:
                button.setProperty("primary", True)
            else:
                button.setProperty("secondary", True)
            button.setMinimumHeight(44)

        layout.addWidget(button_box)

    def get_test_type(self) -> str:
        return self.test_type_input.text().strip()

    def get_description(self) -> str:
        return self.description_input.text().strip()


class SessionState:
    """Minimal session holder for the desktop application."""

    def __init__(self) -> None:
        self.user: Optional[Dict[str, Any]] = None

    def set_user(self, user: Dict[str, Any]) -> None:
        self.user = user

    def clear(self) -> None:
        self.user = None


class LoginPage(QWidget):
    """Modern authentication interface with industrial design."""

    login_requested = pyqtSignal(str, str)
    signup_requested = pyqtSignal(str, str, str)
    forgot_password_requested = pyqtSignal()

    def __init__(self) -> None:
        super().__init__()

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setAlignment(Qt.AlignCenter)
        layout.setContentsMargins(40, 60, 40, 60)
        layout.setSpacing(32)

        # Header
        header_widget = QWidget()
        header_layout = QVBoxLayout(header_widget)
        header_layout.setSpacing(8)
        header_layout.setAlignment(Qt.AlignCenter)

        title = QLabel("Industrial Data System")
        title.setProperty("heading", True)
        header_layout.addWidget(title, alignment=Qt.AlignCenter)

        subtitle = QLabel("Secure data management platform")
        subtitle.setProperty("caption", True)
        header_layout.addWidget(subtitle, alignment=Qt.AlignCenter)

        layout.addWidget(header_widget)

        # Tab Toggle
        toggle_card = QFrame()
        toggle_card.setProperty("card", True)
        toggle_card.setMaximumWidth(480)
        toggle_layout = QVBoxLayout(toggle_card)
        toggle_layout.setContentsMargins(8, 8, 8, 8)
        toggle_layout.setSpacing(0)

        tab_row = QHBoxLayout()
        tab_row.setSpacing(0)
        tab_row.setContentsMargins(0, 0, 0, 0)

        self.login_toggle = QPushButton("Sign In")
        self.signup_toggle = QPushButton("Sign Up")

        for button in (self.login_toggle, self.signup_toggle):
            button.setCheckable(True)
            button.setCursor(Qt.PointingHandCursor)
            button.setMinimumHeight(48)
            button.setProperty("secondary", True)

        self.login_toggle.clicked.connect(lambda: self._switch_mode("login"))
        self.signup_toggle.clicked.connect(lambda: self._switch_mode("signup"))

        tab_row.addWidget(self.login_toggle)
        tab_row.addWidget(self.signup_toggle)
        toggle_layout.addLayout(tab_row)

        layout.addWidget(toggle_card, alignment=Qt.AlignCenter)

        # Form Container
        form_card = QFrame()
        form_card.setProperty("card", True)
        form_card.setMaximumWidth(480)
        form_layout = QVBoxLayout(form_card)
        form_layout.setSpacing(24)
        form_layout.setContentsMargins(32, 32, 32, 32)

        self.form_stack = QStackedWidget()
        form_layout.addWidget(self.form_stack)

        # Login Form
        self.login_form = QWidget()
        login_layout = QVBoxLayout(self.login_form)
        login_layout.setSpacing(20)
        login_layout.setContentsMargins(0, 0, 0, 0)

        login_email_label = QLabel("Email or Username")
        login_email_label.setStyleSheet(f"color: {IndustrialTheme.TEXT_SECONDARY}; font-weight: 500;")
        self.login_email_input = QLineEdit()
        self.login_email_input.setPlaceholderText("Enter your email or username")
        login_layout.addWidget(login_email_label)
        login_layout.addWidget(self.login_email_input)

        login_password_label = QLabel("Password")
        login_password_label.setStyleSheet(f"color: {IndustrialTheme.TEXT_SECONDARY}; font-weight: 500;")
        self.login_password_input = QLineEdit()
        self.login_password_input.setPlaceholderText("Enter your password")
        self.login_password_input.setEchoMode(QLineEdit.Password)
        login_layout.addWidget(login_password_label)
        login_layout.addWidget(self.login_password_input)

        login_button = QPushButton("Sign In")
        login_button.setProperty("primary", True)
        login_button.clicked.connect(self._emit_login_request)
        login_layout.addWidget(login_button)

        forgot_button = QPushButton("Forgot Password?")
        forgot_button.setProperty("flat", True)
        forgot_button.setCursor(Qt.PointingHandCursor)
        forgot_button.clicked.connect(self.forgot_password_requested)
        login_layout.addWidget(forgot_button, alignment=Qt.AlignCenter)

        self.form_stack.addWidget(self.login_form)

        # Signup Form
        self.signup_form = QWidget()
        signup_layout = QVBoxLayout(self.signup_form)
        signup_layout.setSpacing(20)
        signup_layout.setContentsMargins(0, 0, 0, 0)

        signup_email_label = QLabel("Email")
        signup_email_label.setStyleSheet(f"color: {IndustrialTheme.TEXT_SECONDARY}; font-weight: 500;")
        self.signup_email_input = QLineEdit()
        self.signup_email_input.setPlaceholderText("your.email@company.com")
        signup_layout.addWidget(signup_email_label)
        signup_layout.addWidget(self.signup_email_input)

        signup_username_label = QLabel("Username")
        signup_username_label.setStyleSheet(f"color: {IndustrialTheme.TEXT_SECONDARY}; font-weight: 500;")
        self.signup_username_input = QLineEdit()
        self.signup_username_input.setPlaceholderText("Username (max 6 characters)")
        self.signup_username_input.setMaxLength(6)
        signup_layout.addWidget(signup_username_label)
        signup_layout.addWidget(self.signup_username_input)

        signup_password_label = QLabel("Password")
        signup_password_label.setStyleSheet(f"color: {IndustrialTheme.TEXT_SECONDARY}; font-weight: 500;")
        self.signup_password_input = QLineEdit()
        self.signup_password_input.setPlaceholderText("Password (max 6 characters)")
        self.signup_password_input.setEchoMode(QLineEdit.Password)
        self.signup_password_input.setMaxLength(6)
        signup_layout.addWidget(signup_password_label)
        signup_layout.addWidget(self.signup_password_input)

        signup_confirm_label = QLabel("Confirm Password")
        signup_confirm_label.setStyleSheet(f"color: {IndustrialTheme.TEXT_SECONDARY}; font-weight: 500;")
        self.signup_confirm_input = QLineEdit()
        self.signup_confirm_input.setPlaceholderText("Re-enter your password")
        self.signup_confirm_input.setEchoMode(QLineEdit.Password)
        self.signup_confirm_input.setMaxLength(6)
        signup_layout.addWidget(signup_confirm_label)
        signup_layout.addWidget(self.signup_confirm_input)

        signup_button = QPushButton("Create Account")
        signup_button.setProperty("primary", True)
        signup_button.clicked.connect(self._emit_signup_request)
        signup_layout.addWidget(signup_button)

        self.form_stack.addWidget(self.signup_form)

        layout.addWidget(form_card, alignment=Qt.AlignCenter)
        layout.addStretch()

        scroll.setWidget(container)
        main_layout.addWidget(scroll)

        self._switch_mode("login")

    def show_login(self, email: str = "") -> None:
        self.login_email_input.setText(email)
        self.login_password_input.clear()
        self.signup_email_input.clear()
        self.signup_username_input.clear()
        self.signup_password_input.clear()
        self.signup_confirm_input.clear()
        self._switch_mode("login")
        if email:
            self.login_password_input.setFocus()
        else:
            self.login_email_input.setFocus()

    def show_signup(self) -> None:
        self.login_email_input.clear()
        self.login_password_input.clear()
        self._switch_mode("signup")
        self.signup_email_input.setFocus()

    def _switch_mode(self, mode: str) -> None:
        if mode == "signup":
            self.form_stack.setCurrentWidget(self.signup_form)
            self.signup_toggle.setChecked(True)
            self.login_toggle.setChecked(False)
            self.signup_toggle.setProperty("primary", True)
            self.signup_toggle.setProperty("secondary", False)
            self.login_toggle.setProperty("primary", False)
            self.login_toggle.setProperty("secondary", True)
        else:
            self.form_stack.setCurrentWidget(self.login_form)
            self.login_toggle.setChecked(True)
            self.signup_toggle.setChecked(False)
            self.login_toggle.setProperty("primary", True)
            self.login_toggle.setProperty("secondary", False)
            self.signup_toggle.setProperty("primary", False)
            self.signup_toggle.setProperty("secondary", True)

        self.login_toggle.style().unpolish(self.login_toggle)
        self.login_toggle.style().polish(self.login_toggle)
        self.signup_toggle.style().unpolish(self.signup_toggle)
        self.signup_toggle.style().polish(self.signup_toggle)

    def _emit_login_request(self) -> None:
        identifier = self.login_email_input.text().strip()
        password = self.login_password_input.text()
        if not identifier or not password:
            QMessageBox.warning(self, "Industrial Data System", "Username/Email and password are required.")
            return
        self.login_requested.emit(identifier, password)

    def _emit_signup_request(self) -> None:
        email = self.signup_email_input.text().strip()
        username = self.signup_username_input.text().strip()
        password = self.signup_password_input.text()
        confirm = self.signup_confirm_input.text()

        if not email or not username or not password or not confirm:
            QMessageBox.warning(self, "Industrial Data System", "All signup fields are required.")
            return

        if password != confirm:
            QMessageBox.warning(self, "Industrial Data System", "Passwords do not match.")
            return

        self.signup_requested.emit(email, username, password)


class ForgotPasswordPage(QWidget):
    """Password reset interface with modern design."""

    reset_requested = pyqtSignal(str)
    back_requested = pyqtSignal()

    def __init__(self) -> None:
        super().__init__()

        main_layout = QVBoxLayout(self)
        main_layout.setAlignment(Qt.AlignCenter)
        main_layout.setContentsMargins(40, 60, 40, 60)
        main_layout.setSpacing(32)

        # Header
        title = QLabel("Reset Password")
        title.setProperty("heading", True)
        main_layout.addWidget(title, alignment=Qt.AlignCenter)

        subtitle = QLabel("Enter your email to receive reset instructions")
        subtitle.setProperty("caption", True)
        main_layout.addWidget(subtitle, alignment=Qt.AlignCenter)

        # Form Card
        form_card = QFrame()
        form_card.setProperty("card", True)
        form_card.setMaximumWidth(480)
        form_layout = QVBoxLayout(form_card)
        form_layout.setSpacing(20)
        form_layout.setContentsMargins(32, 32, 32, 32)

        email_label = QLabel("Email Address")
        email_label.setStyleSheet(f"color: {IndustrialTheme.TEXT_SECONDARY}; font-weight: 500;")
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("your.email@company.com")
        form_layout.addWidget(email_label)
        form_layout.addWidget(self.email_input)

        reset_button = QPushButton("Send Reset Instructions")
        reset_button.setProperty("primary", True)
        reset_button.clicked.connect(self._emit_reset_request)
        form_layout.addWidget(reset_button)

        back_button = QPushButton("Back to Sign In")
        back_button.setProperty("flat", True)
        back_button.clicked.connect(self.back_requested)
        form_layout.addWidget(back_button, alignment=Qt.AlignCenter)

        main_layout.addWidget(form_card, alignment=Qt.AlignCenter)
        main_layout.addStretch()

    def _emit_reset_request(self) -> None:
        email = self.email_input.text().strip().lower()
        self.reset_requested.emit(email)


class DashboardPage(QWidget):
    """Modern dashboard with test type organization."""

    logout_requested = pyqtSignal()
    upload_requested = pyqtSignal(list, str)  # file_path, test_type
    refresh_requested = pyqtSignal()
    test_type_created = pyqtSignal(str, str)

    def __init__(self) -> None:
        super().__init__()

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(40, 32, 40, 32)
        layout.setSpacing(24)

        # Header Section
        header_layout = QHBoxLayout()
        header_widget = QWidget()
        header_inner = QVBoxLayout(header_widget)
        header_inner.setSpacing(4)
        header_inner.setContentsMargins(0, 0, 0, 0)

        self.welcome_label = QLabel("Dashboard")
        self.welcome_label.setProperty("heading", True)
        header_inner.addWidget(self.welcome_label)

        self.subtitle_label = QLabel("Manage your industrial test data")
        self.subtitle_label.setProperty("caption", True)
        header_inner.addWidget(self.subtitle_label)

        header_layout.addWidget(header_widget)
        header_layout.addStretch()

        # Action Buttons
        action_layout = QHBoxLayout()
        action_layout.setSpacing(12)

        upload_button = QPushButton("Upload File")
        upload_button.setProperty("primary", True)
        upload_button.clicked.connect(self._select_file)
        action_layout.addWidget(upload_button)

        refresh_button = QPushButton("Refresh")
        refresh_button.setProperty("secondary", True)
        refresh_button.clicked.connect(self.refresh_requested.emit)
        action_layout.addWidget(refresh_button)

        logout_button = QPushButton("Sign Out")
        logout_button.setProperty("danger", True)
        logout_button.clicked.connect(self.logout_requested)
        action_layout.addWidget(logout_button)

        header_layout.addLayout(action_layout)
        layout.addLayout(header_layout)

        # Upload Section Card
        upload_card = QFrame()
        upload_card.setProperty("card", True)
        upload_layout = QVBoxLayout(upload_card)
        upload_layout.setSpacing(16)

        upload_title = QLabel("Test Type Selection")
        upload_title.setProperty("subheading", True)
        upload_layout.addWidget(upload_title)

        # Test type selector
        test_type_layout = QHBoxLayout()
        test_type_layout.setSpacing(12)

        test_type_label = QLabel("Test Type:")
        test_type_label.setStyleSheet(f"color: {IndustrialTheme.TEXT_SECONDARY}; font-weight: 500;")
        test_type_layout.addWidget(test_type_label)

        self.test_type_combo = QComboBox()
        self.test_type_combo.setMinimumWidth(250)
        test_type_layout.addWidget(self.test_type_combo, stretch=1)

        new_type_button = QPushButton("+ New Test Type")
        new_type_button.setProperty("secondary", True)
        new_type_button.clicked.connect(self._create_new_test_type)
        test_type_layout.addWidget(new_type_button)

        upload_layout.addLayout(test_type_layout)

        info_label = QLabel(
            "Select a test type before uploading files. Files are saved to the shared drive in folders by test type."
        )
        info_label.setProperty("caption", True)
        info_label.setWordWrap(True)
        upload_layout.addWidget(info_label)

        layout.addWidget(upload_card)

        # Files Card
        files_card = QFrame()
        files_card.setProperty("card", True)
        files_layout = QVBoxLayout(files_card)
        files_layout.setContentsMargins(0, 0, 0, 0)
        files_layout.setSpacing(0)

        files_header = QLabel("Uploaded Files")
        files_header.setProperty("subheading", True)
        files_header.setStyleSheet(f"padding: 20px 24px; border-bottom: 1px solid {IndustrialTheme.BORDER};")
        files_layout.addWidget(files_header)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Filename", "Test Type", "Path", "Uploaded"])
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setFocusPolicy(Qt.NoFocus)
        files_layout.addWidget(self.table)

        actions_layout = QHBoxLayout()
        actions_layout.setContentsMargins(16, 12, 16, 16)
        actions_layout.setSpacing(12)

        self.open_file_button = QPushButton("Open File")
        self.open_file_button.setProperty("secondary", True)
        self.open_file_button.clicked.connect(self._open_selected_file)
        actions_layout.addWidget(self.open_file_button)

        self.open_folder_button = QPushButton("Open in Explorer")
        self.open_folder_button.setProperty("secondary", True)
        self.open_folder_button.clicked.connect(self._open_selected_folder)
        actions_layout.addWidget(self.open_folder_button)

        self.copy_path_button = QPushButton("Copy Path")
        self.copy_path_button.setProperty("secondary", True)
        self.copy_path_button.clicked.connect(self._copy_selected_path)
        actions_layout.addWidget(self.copy_path_button)

        self.show_properties_button = QPushButton("Show Properties")
        self.show_properties_button.setProperty("secondary", True)
        self.show_properties_button.clicked.connect(self._show_selected_properties)
        actions_layout.addWidget(self.show_properties_button)

        actions_layout.addStretch()
        files_layout.addLayout(actions_layout)

        layout.addWidget(files_card)

        # CSV Preview Card
        self.csv_card = QFrame()
        self.csv_card.setProperty("card", True)
        csv_layout = QVBoxLayout(self.csv_card)
        csv_layout.setContentsMargins(0, 0, 0, 0)
        csv_layout.setSpacing(0)

        self.csv_preview_label = QLabel("CSV Preview")
        self.csv_preview_label.setProperty("subheading", True)
        self.csv_preview_label.setStyleSheet(f"padding: 20px 24px; border-bottom: 1px solid {IndustrialTheme.BORDER};")
        csv_layout.addWidget(self.csv_preview_label)

        self.csv_table = QTableWidget(0, 0)
        self.csv_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.csv_table.verticalHeader().setVisible(False)
        self.csv_table.setAlternatingRowColors(True)
        csv_layout.addWidget(self.csv_table)

        self.csv_card.hide()
        layout.addWidget(self.csv_card)

        scroll.setWidget(container)
        main_layout.addWidget(scroll)

        self.test_types: List[str] = []

    def set_test_types(self, test_types: List[str]) -> None:
        """Update the test type dropdown with available types."""
        self.test_types = sorted(test_types)
        self.test_type_combo.clear()
        if self.test_types:
            self.test_type_combo.addItems(self.test_types)
        else:
            self.test_type_combo.addItem("No test types available")

    def get_selected_test_type(self) -> Optional[str]:
        """Get the currently selected test type."""
        if not self.test_types:
            return None
        return self.test_type_combo.currentText()

    def _create_new_test_type(self) -> None:
        """Show dialog to create a new test type."""
        dialog = NewTestTypeDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            test_type = dialog.get_test_type()
            description = dialog.get_description()
            if test_type:
                self.test_type_created.emit(test_type, description)
                # Add to combo box
                if test_type not in self.test_types:
                    self.test_types.append(test_type)
                    self.test_types.sort()
                    self.test_type_combo.clear()
                    self.test_type_combo.addItems(self.test_types)
                    # Select the new type
                    index = self.test_type_combo.findText(test_type)
                    if index >= 0:
                        self.test_type_combo.setCurrentIndex(index)
                else:
                    QMessageBox.information(self, "Industrial Data System", f"Test type '{test_type}' already exists.")

    def set_user_identity(self, username: str, email: str) -> None:
        username = username.strip()
        email = email.strip()
        if username:
            self.welcome_label.setText(f"Welcome, {username}")
            self.subtitle_label.setText(email if email else "Manage your industrial test data")
        elif email:
            self.welcome_label.setText(f"Welcome, {email}")
            self.subtitle_label.setText("Manage your industrial test data")
        else:
            self.welcome_label.setText("Dashboard")
            self.subtitle_label.setText("Manage your industrial test data")

    def update_files(self, files: List[Dict[str, Any]]) -> None:
        self.file_records = files
        self.table.setRowCount(len(files))
        for row, file_record in enumerate(files):
            filename_item = QTableWidgetItem(file_record.get("filename", ""))
            test_type_item = QTableWidgetItem(file_record.get("test_type", ""))
            path_item = QTableWidgetItem(file_record.get("absolute_path") or file_record.get("file_path", ""))
            created_item = QTableWidgetItem(file_record.get("created_at", ""))

            for item in (filename_item, test_type_item, path_item, created_item):
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)

            self.table.setItem(row, 0, filename_item)
            self.table.setItem(row, 1, test_type_item)
            self.table.setItem(row, 2, path_item)
            self.table.setItem(row, 3, created_item)

    def _get_selected_record(self) -> Optional[Dict[str, Any]]:
        selection = self.table.selectionModel().selectedRows()
        if not selection:
            return None
        row = selection[0].row()
        if 0 <= row < len(self.file_records):
            return self.file_records[row]
        return None

    def _open_selected_file(self) -> None:
        record = self._get_selected_record()
        if not record:
            QMessageBox.information(self, "Industrial Data System", "Select a file first.")
            return
        path_value = record.get("absolute_path") or record.get("file_path")
        if not path_value:
            QMessageBox.warning(self, "Industrial Data System", "No file path available.")
            return
        path = Path(path_value)
        if not path.is_absolute():
            base_path = record.get("base_path")
            if base_path:
                path = Path(base_path) / path
        if not path.exists():
            QMessageBox.warning(self, "Industrial Data System", f"File not found: {path}")
            return
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(path)))

    def _open_selected_folder(self) -> None:
        record = self._get_selected_record()
        if not record:
            QMessageBox.information(self, "Industrial Data System", "Select a file first.")
            return
        path_value = record.get("absolute_path") or record.get("file_path")
        if not path_value:
            QMessageBox.warning(self, "Industrial Data System", "No file path available.")
            return
        path = Path(path_value)
        if not path.is_absolute():
            base_path = record.get("base_path")
            if base_path:
                path = Path(base_path) / path
        if not path.exists():
            QMessageBox.warning(self, "Industrial Data System", f"File not found: {path}")
            return
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(path.parent)))

    def _copy_selected_path(self) -> None:
        record = self._get_selected_record()
        if not record:
            QMessageBox.information(self, "Industrial Data System", "Select a file first.")
            return
        path_value = record.get("absolute_path") or record.get("file_path")
        if not path_value:
            QMessageBox.warning(self, "Industrial Data System", "No file path available.")
            return
        path = Path(path_value)
        if not path.is_absolute():
            base_path = record.get("base_path")
            if base_path:
                path = Path(base_path) / path
        QApplication.clipboard().setText(str(path))
        QMessageBox.information(self, "Industrial Data System", "File path copied to clipboard.")

    def _show_selected_properties(self) -> None:
        record = self._get_selected_record()
        if not record:
            QMessageBox.information(self, "Industrial Data System", "Select a file first.")
            return
        path_value = record.get("absolute_path") or record.get("file_path")
        if not path_value:
            QMessageBox.warning(self, "Industrial Data System", "No file path available.")
            return
        path = Path(path_value)
        if not path.is_absolute():
            base_path = record.get("base_path")
            if base_path:
                path = Path(base_path) / path
        if not path.exists():
            QMessageBox.warning(self, "Industrial Data System", f"File not found: {path}")
            return
        stat = path.stat()
        size_mb = stat.st_size / (1024 * 1024)
        modified = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        created = datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d %H:%M:%S")
        QMessageBox.information(
            self,
            "Industrial Data System",
            (
                f"Location: {path}\n"
                f"Size: {size_mb:.2f} MB\n"
                f"Created: {created}\n"
                f"Modified: {modified}"
            ),
        )

    def _select_file(self) -> None:
        test_type = self.get_selected_test_type()
        if not test_type or test_type == "No test types available":
            QMessageBox.warning(
                self,
                "Industrial Data System",
                "Please select or create a test type before uploading files."
            )
            return

        file_dialog_filter = "Data Files (*.csv *.xlsx *.xlsm *.xltx *.xltm *.asc);;"
        file_dialog_filter += "CSV Files (*.csv);;ASC Files (*.asc);;Excel Files (*.xlsx *.xlsm *.xltx *.xltm)"

        # Changed from getOpenFileName to getOpenFileNames (plural)
        file_paths, _ = QFileDialog.getOpenFileNames(  # <-- THIS IS THE KEY CHANGE
            self,
            "Select Files to Upload",  # Updated text
            "",
            file_dialog_filter,
        )
        if file_paths:  # This is now a list
            self.upload_requested.emit(file_paths, test_type)  # Send list instead of single path

    def display_csv_preview(self, headers: List[str], rows: List[List[str]]) -> None:
        if not headers:
            self.clear_csv_preview()
            return

        column_count = len(headers)
        self.csv_table.setColumnCount(column_count)
        self.csv_table.setHorizontalHeaderLabels(headers)

        self.csv_table.setRowCount(len(rows))
        for row_index, row_values in enumerate(rows):
            for column_index in range(column_count):
                value = row_values[column_index] if column_index < len(row_values) else ""
                item = QTableWidgetItem(value)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.csv_table.setItem(row_index, column_index, item)

        self.csv_card.show()

    def clear_csv_preview(self) -> None:
        self.csv_table.clear()
        self.csv_table.setRowCount(0)
        self.csv_table.setColumnCount(0)
        self.csv_card.hide()


class IndustrialDataApp(QMainWindow):
    """Main window with modern industrial UI design."""

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Industrial Data System")
        self.resize(1200, 800)
        self.setMinimumSize(900, 600)

        self.session_state = SessionState()
        self.db_manager = DatabaseManager()
        self.auth_store = LocalAuthStore(self.db_manager)
        self.history_store = UploadHistoryStore(self.db_manager)
        self.storage_manager = LocalStorageManager(config=CONFIG, database=self.db_manager)
        self.current_username: str = ""

        self.stack = QStackedWidget()
        self.setCentralWidget(self.stack)

        self.login_page = LoginPage()
        self.forgot_page = ForgotPasswordPage()
        self.dashboard_page = DashboardPage()

        self.stack.addWidget(self.login_page)
        self.stack.addWidget(self.forgot_page)
        self.stack.addWidget(self.dashboard_page)

        self.login_page.login_requested.connect(self.handle_login)
        self.login_page.signup_requested.connect(self.handle_signup)
        self.login_page.forgot_password_requested.connect(self.show_forgot_password)

        self.forgot_page.reset_requested.connect(self.handle_password_reset)
        self.forgot_page.back_requested.connect(self.show_login)

        self.dashboard_page.logout_requested.connect(self.handle_logout)
        self.dashboard_page.upload_requested.connect(self.handle_upload)
        self.dashboard_page.refresh_requested.connect(self.refresh_files)
        self.dashboard_page.test_type_created.connect(self.handle_new_test_type)

        self.show_login()

    def show_login(self, email: str = "") -> None:
        self.stack.setCurrentWidget(self.login_page)
        self.login_page.show_login(email)

    def show_forgot_password(self) -> None:
        self.stack.setCurrentWidget(self.forgot_page)

    def show_dashboard(self) -> None:
        self.stack.setCurrentWidget(self.dashboard_page)
        self.load_test_types()

    def load_test_types(self) -> None:
        """Load available test types from the database and shared drive."""
        try:
            records = self.db_manager.list_test_types()
            test_types = {record.name for record in records}
            tests_dir = CONFIG.files_base_path / "tests"
            if tests_dir.exists():
                for child in tests_dir.iterdir():
                    if child.is_dir():
                        test_types.add(child.name)
            test_type_list = sorted(test_types)
            self.dashboard_page.set_test_types(test_type_list)
        except Exception as exc:
            self.dashboard_page.set_test_types([])

    def handle_new_test_type(self, name: str, description: str) -> None:
        description_value = description.strip() or None
        try:
            record = self.db_manager.ensure_test_type(name, description_value)
            self.storage_manager.ensure_folder_exists(record.name)
        except StorageError as exc:
            self._alert(str(exc), QMessageBox.Warning)
            return
        except Exception as exc:
            self._alert(f"Unable to create test type: {exc}", QMessageBox.Critical)
            return
        self.load_test_types()
        index = self.dashboard_page.test_type_combo.findText(record.name)
        if index >= 0:
            self.dashboard_page.test_type_combo.setCurrentIndex(index)

    def handle_login(self, identifier: str, password: str) -> None:
        identifier = identifier.strip()
        if not identifier or not password:
            self._alert("Username/Email and password are required.", QMessageBox.Warning)
            return

        user = self.auth_store.authenticate(identifier, password)
        if not user:
            self._alert("Invalid credentials. Please try again.", QMessageBox.Warning)
            return

        self._set_logged_in_user(user)

    def _set_logged_in_user(self, user: LocalUser) -> None:
        session_payload = {
            "id": user.id,
            "email": user.email,
            "username": (
                user.username
                or user.metadata.get("username")
                or user.metadata.get("username_normalized")
                or ""
            ),
            "metadata": user.metadata,
        }
        self.session_state.set_user(session_payload)
        self.current_username = session_payload["username"] or ""
        self.dashboard_page.set_user_identity(
            self.current_username, session_payload.get("email", "")
        )
        self.show_dashboard()
        self.refresh_files()

    def handle_signup(self, email: str, username: str, password: str) -> None:
        email = email.strip().lower()
        username = username.strip()
        if not email or not username or not password:
            self._alert("All signup fields are required.", QMessageBox.Warning)
            return

        if len(username) > 6 or len(password) > 6:
            self._alert(
                "Username and password must be 6 characters or fewer.",
                QMessageBox.Warning,
            )
            return

        if not self._is_valid_email(email):
            self._alert("Enter a valid email address.", QMessageBox.Warning)
            return

        normalized_username = username.lower()

        metadata = {
            "username": username,
            "username_normalized": normalized_username,
        }

        try:
            user = self.auth_store.create_user(
                email=email,
                password=password,
                username=normalized_username,
                metadata=metadata,
            )
        except ValueError as exc:
            self._alert(str(exc), QMessageBox.Warning)
            return

        self._alert(
            "Signup successful! You're now signed in.",
            QMessageBox.Information,
        )
        self._set_logged_in_user(user)

    def handle_password_reset(self, email: str) -> None:
        if not email:
            self._alert("Email is required to reset the password.", QMessageBox.Warning)
            return

        user_exists = any(user.email == email.strip().lower() for user in self.auth_store.list_users())
        if not user_exists:
            self._alert("No account was found with that email.", QMessageBox.Warning)
            return

        self._alert(
            "Password resets must be handled by an administrator for local accounts.",
            QMessageBox.Information,
        )
        self.show_login(email)

    def handle_logout(self) -> None:
        self.session_state.clear()
        self.current_username = ""
        self._alert("You have been signed out.", QMessageBox.Information)
        self.show_login()

    def refresh_files(self) -> None:
        user = self.session_state.user
        if not user:
            self._alert("Session expired. Please sign in again.", QMessageBox.Warning)
            self.show_login()
            return

        metadata = user.get("metadata") or {}
        username = (
            user.get("username")
            or metadata.get("username")
            or metadata.get("username_normalized")
            or self.current_username
            or ""
        ).strip()
        self.current_username = username

        self.dashboard_page.set_user_identity(
            self.current_username,
            user.get("email", ""),
        )

        user_id = user.get("id")
        if user_id is None:
            self._alert("Session error: missing user identifier.", QMessageBox.Warning)
            self.show_login()
            return

        base_path = self.storage_manager.base_path
        if base_path.exists():
            self.db_manager.prune_missing_uploads(base_path)

        records = []
        for record in self.history_store.get_records_for_user(int(user_id)):
            relative_path = record.get("file_path")
            absolute_path = None
            if relative_path:
                absolute_candidate = (CONFIG.files_base_path / Path(relative_path)).resolve()
                absolute_path = str(absolute_candidate)
            record["absolute_path"] = absolute_path
            record["base_path"] = str(CONFIG.files_base_path)
            records.append(record)

        self.dashboard_page.update_files(records)
        self.load_test_types()

    def handle_upload(self, file_paths: str | List[str], test_type: str) -> None:
        # Convert single file to list for uniform handling
        if isinstance(file_paths, str):
            file_paths = [file_paths]

        user = self.session_state.user
        if not user:
            self._alert("Session expired. Please sign in again.", QMessageBox.Warning)
            self.show_login()
            return

        if not test_type:
            self._alert("Please select a test type.", QMessageBox.Warning)
            return

        supported_extensions = {".csv", ".xlsx", ".xlsm", ".xltx", ".xltm", ".asc"}

        successful_uploads = []
        failed_uploads = []

        for file_path in file_paths:
            file_extension = Path(file_path).suffix.lower()

            if file_extension not in supported_extensions:
                allowed = ", ".join(sorted(supported_extensions))
                failed_uploads.append((file_path, f"Unsupported file type. Allowed: {allowed}"))
                continue

            # Only show preview for single file
            if len(file_paths) == 1:
                preview_result = self._prepare_file_preview(file_path, file_extension)
                if preview_result is None:
                    failed_uploads.append((file_path, "Preview generation failed"))
                    continue
                headers, rows = preview_result
                self.dashboard_page.display_csv_preview(headers, rows)
            else:
                self.dashboard_page.clear_csv_preview()

            try:
                stored = self.storage_manager.upload_file(file_path, test_type)
            except StorageError as exc:
                failed_uploads.append((file_path, str(exc)))
                continue

            try:
                self.history_store.add_record(
                    user_id=int(user.get("id")),
                    filename=os.path.basename(file_path),
                    file_path=str(stored.relative_path),
                    test_type=test_type,
                    file_size=stored.size_bytes,
                )
                successful_uploads.append(file_path)
            except Exception as exc:
                # Attempt to clean up the copied file if database write fails
                try:
                    self.storage_manager.delete_file(stored.absolute_path)
                except StorageError:
                    pass
                failed_uploads.append((file_path, f"Failed to record upload: {exc}"))
                continue

        # Show summary message
        if len(file_paths) > 1:
            summary = f"Upload Complete:\n"
            summary += f"✓ Successfully uploaded: {len(successful_uploads)} files\n"
            if failed_uploads:
                summary += f"✗ Failed: {len(failed_uploads)} files\n\n"
                summary += "Failed files:\n"
                for fname, error in failed_uploads[:5]:  # Show first 5 failures
                    summary += f"  - {Path(fname).name}: {error}\n"
                if len(failed_uploads) > 5:
                    summary += f"  ... and {len(failed_uploads) - 5} more\n"
            self._alert(summary, QMessageBox.Information if successful_uploads else QMessageBox.Warning)
        elif successful_uploads:
            # Single file success message
            stored_path = self.storage_manager.get_file_path(test_type, os.path.basename(successful_uploads[0]))
            self._alert(
                f"File uploaded to shared drive at: {stored_path}",
                QMessageBox.Information,
            )
        elif failed_uploads:
            # Single file failure
            self._alert(failed_uploads[0][1], QMessageBox.Critical)

        if successful_uploads:
            self.refresh_files()

    @staticmethod
    def _is_valid_email(email: str) -> bool:
        return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email))

    def _prepare_file_preview(
            self, file_path: str, file_extension: str
    ) -> Optional[tuple[List[str], List[List[str]]]]:
        rows: List[List[str]] = []
        if file_extension == ".csv" or file_extension == ".asc":
            try:
                # Try UTF-8 first, fall back to latin-1 or cp1252
                encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                for encoding in encodings:
                    try:
                        with open(file_path, newline="", encoding=encoding) as file:
                            reader = csv.reader(file)
                            rows = list(reader)
                        break  # Success, exit loop
                    except UnicodeDecodeError:
                        if encoding == encodings[-1]:  # Last encoding failed
                            raise
                        continue  # Try next encoding
            except Exception as exc:
                self._alert(f"Unable to read file: {exc}", QMessageBox.Critical)
                self.dashboard_page.clear_csv_preview()
                return None
        else:

            try:
                from openpyxl import load_workbook
            except ImportError:
                self._alert(
                    "Excel support is unavailable because openpyxl is not installed.",
                    QMessageBox.Critical,
                )
                self.dashboard_page.clear_csv_preview()
                return None
            workbook = None
            try:
                workbook = load_workbook(
                    filename=file_path,
                    read_only=True,
                    data_only=True,
                )
                worksheet = workbook.active
                for row in worksheet.iter_rows(values_only=True):
                    rows.append([
                        "" if cell is None else str(cell)
                        for cell in row
                    ])
                    if len(rows) >= 101:
                        break
            except Exception as exc:
                self._alert(
                    f"Unable to read Excel file: {exc}", QMessageBox.Critical,
                )
                self.dashboard_page.clear_csv_preview()
                return None
            finally:
                if workbook is not None:
                    try:
                        workbook.close()
                    except Exception:
                        pass

        if not rows:
            self._alert("The selected file is empty.", QMessageBox.Warning)
            self.dashboard_page.clear_csv_preview()
            return None

        headers = [str(value) for value in rows[0]] if rows else []
        data_rows = [
            [str(value) for value in row]
            for row in rows[1:101]
        ]
        return headers, data_rows

    def _alert(self, message: str, icon: QMessageBox.Icon) -> None:
        dialog = QMessageBox(self)
        dialog.setIcon(icon)
        dialog.setText(message)
        dialog.setWindowTitle("Industrial Data System")
        dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {IndustrialTheme.SURFACE};
            }}
            QMessageBox QLabel {{
                color: {IndustrialTheme.TEXT_PRIMARY};
                font-size: 14px;
                padding: 10px;
            }}
            QPushButton {{
                min-width: 80px;
                padding: 8px 16px;
            }}
        """)
        dialog.exec_()


def main() -> None:
    app = QApplication(sys.argv)

    # Apply industrial theme stylesheet
    app.setStyleSheet(IndustrialTheme.get_stylesheet())

    # Set application-wide font
    app.setFont(QFont("Segoe UI", 10))

    window = IndustrialDataApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
